"""
crystal_pbi_generator.py  —  Crystal Reports to Power BI Migration Server
Runs as Flask web server only.  Usage: python crystal_pbi_generator.py [--port N] [--debug]
"""
import os, sys, re, uuid, json, zipfile, logging, platform, subprocess, threading, traceback, argparse
from pathlib import Path
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Optional

# ── Server version ────────────────────────────────────────────────────────────
# Format: MAJOR.MINOR.WIP
#   MAJOR — breaking structural changes (new PBIT format, major architecture)
#   MINOR — new features (new data source types, new visual types, new sheets)
#   WIP   — bug fixes and incremental improvements
SERVER_VERSION = "1.5.7"

def _pip(pkg):
    subprocess.run([sys.executable,"-m","pip","install",pkg,"--quiet","--break-system-packages"],check=False)
try:    import flask
except: _pip("flask"); import flask
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except:
    _pip("openpyxl"); import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
try:    import olefile
except: _pip("olefile"); import olefile

CRYSTAL_SDK_AVAILABLE = False
CRYSTAL_SDK_MODE = None   # "clr" | "com" | None

if platform.system() == "Windows":
    # Prefer pythonnet (clr) — talks directly to the .NET assemblies,
    # works with SAP Crystal Reports SDK for .NET (VS/runtime installs)
    try:
        import clr as _clr_probe  # noqa: F401
        CRYSTAL_SDK_AVAILABLE = True
        CRYSTAL_SDK_MODE = "clr"
    except ImportError:
        pass

    # Fall back to win32com if pythonnet not installed
    if not CRYSTAL_SDK_AVAILABLE:
        try:
            import win32com.client as win32  # noqa: F401
            CRYSTAL_SDK_AVAILABLE = True
            CRYSTAL_SDK_MODE = "com"
        except ImportError:
            pass

from flask import Flask, request, jsonify, Response, send_from_directory

logging.basicConfig(level=logging.INFO,format="%(asctime)s  %(levelname)-8s  %(message)s",datefmt="%H:%M:%S")
log = logging.getLogger("Crystal2PBI")
_sse_messages: list = []; _sse_lock = threading.Lock()
def emit(msg,level="INFO"):
    (log.info if level=="INFO" else log.warning)(msg)
    with _sse_lock: _sse_messages.append(json.dumps({"level":level,"msg":msg}))

# ── Data models ──────────────────────────────────────────────────────────────
class CrystalField:
    def __init__(self, name, field_type="database", data_type="string",
                 formula_text="", table_name="", source_column=""):
        self.name         = name
        self.field_type   = field_type    # database|formula|parameter|running_total|special
        self.data_type    = data_type     # string|number|currency|date|datetime|boolean
        self.formula_text = formula_text
        self.table_name   = table_name
        self.source_column= source_column or name
        # Formatting hints extracted from report objects
        self.font_name    = "Arial"
        self.font_size    = 10
        self.bold         = False
        self.italic       = False
        self.h_align      = "left"        # left|center|right


class CrystalTable:
    def __init__(self, name, alias="", db_type="odbc", connection_string="",
                 sql_command="", server="", database="", connection_type=""):
        self.name            = name
        self.alias           = alias or name
        self.db_type         = db_type        # odbc|sql|access|excel|unknown|seeded
        self.connection_string = connection_string
        self.sql_command     = sql_command    # CommandTable SQL (non-empty when it's a SQL command)
        self.server          = server
        self.database        = database
        self.connection_type = connection_type  # e.g. "ODBC", "OLE DB", "Native SQL"
        self.fields: List[CrystalField] = []


class CrystalSortField:
    def __init__(self, field_name: str, direction: str = "asc", sort_type: str = "record"):
        self.field_name = field_name    # e.g. "countries_all_iso.name"
        self.direction  = direction     # "asc" | "desc"
        self.sort_type  = sort_type     # "record" | "group"


class CrystalSubreport:
    def __init__(self, name: str, section: str = ""):
        self.name         = name          # subreport report name
        self.section      = section       # parent section it lives in
        self.link_fields: List[dict] = [] # [{main_field, sub_field, link_type}]
        self.x = self.y = self.w = self.h = 0


class CrystalSection:
    def __init__(self, name, section_type="detail"):
        self.name             = name
        self.section_type     = section_type  # reportHeader|pageHeader|groupHeader|
                                              # detail|groupFooter|pageFooter|reportFooter
        self.objects: list    = []
        self.suppress         = False
        self.suppress_formula = ""
        self.height           = 0             # twips (1440 twips = 1 inch)
        self.background_color = ""            # hex e.g. "#FFFFFF"
        self.group_level      = 0             # for groupHeader/Footer: which group level


class CrystalReport:
    def __init__(self):
        self.file_path    = ""
        self.report_name  = ""
        self.report_title = ""
        self.subject      = ""
        self.author       = ""
        self.tables:    List[CrystalTable]   = []
        self.fields:    List[CrystalField]   = []
        self.parameters: list                = []
        self.groups:     list                = []   # [{field, order, group_level}]
        self.sort_fields: List[CrystalSortField] = []
        self.record_selection_formula = ""
        self.group_selection_formula  = ""
        self.sections:   List[CrystalSection]  = []
        self.subreports: List[CrystalSubreport] = []
        self.warnings:   list = []
        self.parse_method = "heuristic"
        # Set by ReportAnalyzer after extraction
        self.recommended_output: str = ""  # "pbit" | "rdl"
        self.routing_reason:     str = ""
        self.routing_scores:     dict = {}

# ═══════════════════════════════════════════════════════════════════════════════
# PBIT SEEDER — reads a reference .pbit to seed table/column schema cross-platform
# ═══════════════════════════════════════════════════════════════════════════════
class PBITSeeder:
    """
    Extracts table name, column definitions, M query, and sort column from an
    existing Power BI .pbit file and injects them into a CrystalReport.

    This is the cross-platform solution for encrypted RPT files: the user
    supplies a reference .pbit (either uploaded via POST /seed or placed
    alongside the .rpt with the same base name) and the seeder fills in
    everything the encrypted RPT binary cannot provide.

    Schema keys returned by extract_schema():
        tables: list of {name, columns:[{name,type,dataType,summarizeBy,
                sourceColumn,expression}], m_expr:[str...], sort_col:str}
    """

    def extract_schema(self, pbit_source) -> Optional[Dict]:
        """Return schema dict or None on failure. pbit_source: path str or bytes."""
        try:
            raw = (open(pbit_source,"rb").read()
                   if isinstance(pbit_source,(str,Path)) else pbit_source)
            with zipfile.ZipFile(BytesIO(raw)) as z:
                dm_bytes = z.read("DataModelSchema")
                # Reference PBIT: pure UTF-16-LE, no BOM.
                # Older generated files: UTF-16-LE with \xff\xfe BOM prefix.
                # Handle both by stripping BOM after decode if present.
                dm_text = dm_bytes.decode("utf-16-le")
                if dm_text.startswith('\ufeff'):
                    dm_text = dm_text[1:]
                schema = json.loads(dm_text)
                layout = json.loads(z.read("Report/Layout"))
        except Exception as e:
            emit(f"  Seeder: cannot read PBIT — {e}", "WARN")
            return None

        tables_out = []
        for tbl in schema["model"]["tables"]:
            cols = []
            for col in tbl.get("columns", []):
                cols.append({
                    "name":         col["name"],
                    "col_type":     col.get("type","regular"),  # regular|calculated
                    "dataType":     col.get("dataType","string"),
                    "summarizeBy":  col.get("summarizeBy","none"),
                    "sourceColumn": col.get("sourceColumn", col["name"]),
                    "expression":   col.get("expression",""),
                })
            m_expr = []
            for part in tbl.get("partitions",[]):
                src = part.get("source",{})
                if src.get("type") == "m":
                    e = src.get("expression",[])
                    m_expr = e if isinstance(e,list) else [e]
                    break
            tables_out.append({
                "name":     tbl["name"],
                "columns":  cols,
                "m_expr":   m_expr,
                "sort_col": self._find_sort_col(layout, tbl["name"]),
            })

        return {"tables": tables_out} if tables_out else None

    @staticmethod
    def _find_sort_col(layout: dict, table_name: str) -> str:
        """Find the sort column by looking at dataTransforms selects for sort=1."""
        try:
            for sec in layout.get("sections",[]):
                for vc in sec.get("visualContainers",[]):
                    if "dataTransforms" not in vc: continue
                    dt = json.loads(vc["dataTransforms"])
                    for sel in dt.get("selects",[]):
                        if sel.get("sort") == 1:
                            qn = sel.get("queryName","")
                            if "." in qn: return qn.split(".")[-1]
        except Exception:
            pass
        return ""

    def apply(self, schema: dict, report: CrystalReport):
        """Merge extracted schema into report, overwriting placeholder data."""
        if not schema or not schema.get("tables"):
            return
        report.tables.clear()
        # Keep only formula fields not associated with a table
        report.fields = [f for f in report.fields if f.field_type == "formula"]

        for st in schema["tables"]:
            tbl = CrystalTable(
                name=st["name"], alias=st["name"], db_type="seeded",
                connection_string="[seeded from reference PBIT]")
            tbl._sort_col = st.get("sort_col","")   # stash for builder
            tbl._m_expr   = st.get("m_expr",[])     # stash for builder

            for col in st["columns"]:
                if col["col_type"] == "regular":
                    f = CrystalField(
                        name=col["name"], field_type="database",
                        data_type=self._pbi2crystal(col["dataType"]),
                        table_name=tbl.name,
                        source_column=col["sourceColumn"])
                    tbl.fields.append(f)
                    report.fields.append(f)
                elif col["col_type"] == "calculated" and col["name"] != "RowColor":
                    # Preserve non-RowColor calculated columns (e.g. CCTLD_formatted)
                    expr = col["expression"]
                    if isinstance(expr, list):
                        expr = "\n".join(str(e) for e in expr if str(e).strip())
                    report.fields.append(CrystalField(
                        name=col["name"], field_type="formula",
                        data_type=self._pbi2crystal(col["dataType"]),
                        formula_text=expr, table_name=tbl.name))

            report.tables.append(tbl)

        report.parse_method = "seeded"
        emit(f"  Seeder: applied schema — {len(report.tables)} table(s), "
             f"{sum(len(t.fields) for t in report.tables)} source columns")

    @staticmethod
    def _pbi2crystal(pbi_type: str) -> str:
        return {"int64":"number","int32":"number","double":"currency",
                "decimal":"currency","string":"string","boolean":"boolean",
                "dateTime":"datetime","date":"date"}.get(pbi_type,"string")


# ── RPT Parser ───────────────────────────────────────────────────────────────
class RPTParser:
    def parse(self, rpt_path: str, seed_pbit: str = None) -> CrystalReport:
        r = CrystalReport(); r.file_path=rpt_path; r.report_name=Path(rpt_path).stem
        emit(f"Parsing: {Path(rpt_path).name}")

        # ── Try SAP Crystal SDK first (Windows only) ──────────────────────────────
        if CRYSTAL_SDK_AVAILABLE:
            try:
                CrystalSDKBridge().load(rpt_path, r)
                r.parse_method = "sdk"
                emit("  ✓ SAP Crystal SDK (COM)")
                return r
            except Exception as e:
                emit(f"  SDK failed ({e}), OLE fallback", "WARN")

        # ── OLE heuristic parse (no SDK available) ────────────────────────────────
        self._parse_ole(rpt_path, r)
        return r

    def _parse_ole(self,path,r):
        try: ole=olefile.OleFileIO(path)
        except Exception as e: r.warnings.append(f"Cannot open OLE: {e}"); self._placeholder(r); return
        self._summary_info(ole,r); self._datasource_stream(ole,r); self._saved_records(ole,r); ole.close()
        if not r.tables: self._placeholder(r)
        if not r.sections:
            for st in ["reportHeader","pageHeader","detail","pageFooter","reportFooter"]:
                r.sections.append(CrystalSection(st,st))

    def _summary_info(self,ole,r):
        try: data=ole.openstream("\x05SummaryInformation").read()
        except: return
        strings,seen=[],set()
        for m in re.finditer(rb"(?:[\x20-\x7e]\x00){4,}",data):
            try:
                s=m.group().decode("utf-16-le").strip()
                if len(s)>=4 and any(c.isalpha() for c in s) and s not in seen: strings.append(s); seen.add(s)
            except: pass
        if strings: r.report_title=strings[0]
        if len(strings)>1: r.subject=strings[1]
        if len(strings)>2 and "SAP" not in strings[2] and "Crystal" not in strings[2]: r.author=strings[2]

    def _datasource_stream(self,ole,r):
        try: data=ole.openstream("DataSourceManager 156l").read()
        except: return
        for m in re.finditer(rb"(?:[\x20-\x7e]\x00){5,}",data):
            try:
                s=m.group().decode("utf-16-le").strip(); lo=s.lower()
                if any(x in lo for x in [".mdb",".accdb",".xls",".db"]):
                    db="access" if ".mdb" in lo or ".accdb" in lo else "file"
                    r.tables.append(CrystalTable(Path(s).stem,db_type=db,connection_string=s)); return
                if "dsn=" in lo or "driver=" in lo:
                    r.tables.append(CrystalTable(s.split("=")[-1].strip(";"),db_type="odbc",connection_string=s)); return
            except: pass

    def _saved_records(self,ole,r):
        try: data=ole.openstream("SavedRecordsStream 158l").read()
        except: return
        if not r.tables: return
        tbl=r.tables[0]
        for m in re.finditer(rb"(?:[\x20-\x7e]\x00){3,50}",data):
            try:
                s=m.group().decode("utf-16-le").strip()
                if re.match(r"^[A-Za-z][A-Za-z0-9_]{2,}$",s):
                    f=CrystalField(s,"database",self._gtype(s),table_name=tbl.name); tbl.fields.append(f); r.fields.append(f)
            except: pass

    def _placeholder(self,r):
        r.tables.append(CrystalTable(r.report_name,db_type="unknown"))
        r.warnings.append(
            "RPT binary is encrypted — no datasource/columns extracted. "
            "To fix: (a) run on Windows with SAP Crystal Runtime, OR "
            "(b) place a reference .pbit with the same base name alongside the .rpt."
        )

    @staticmethod
    def _gtype(col):
        n=col.lower()
        if any(k in n for k in ["id","num","count","total","amount","qty","price"]): return "number"
        if any(k in n for k in ["date","time","created","modified"]): return "datetime"
        if any(k in n for k in ["flag","active","enabled"]): return "boolean"
        return "string"

# ── Crystal SDK Bridge ────────────────────────────────────────────────────────
# Supports two loading paths (tried in this order):
#
#   1. pythonnet / clr  — SAP Crystal Reports SDK for .NET (preferred)
#      pip install pythonnet
#      Works with the VS developer package or the standalone .NET runtime.
#      Set env var CRYSTAL_ASSEMBLY_PATH if assemblies are in a non-standard
#      location (the dir that contains CrystalDecisions.CrystalReports.Engine.dll).
#
#   2. win32com — legacy COM / ActiveX runtime
#      Tries several versioned ProgIDs in descending order.
#      NOTE: the .NET SDK does NOT register COM ProgIDs, so if you installed
#      "SAP Crystal Reports, Developer Version for Visual Studio" you will get
#      error -2147221005 (REGDB_E_CLASSNOTREG) with win32com.
#      Solution: pip install pythonnet  (or use a reference PBIT for seeding).
# ─────────────────────────────────────────────────────────────────────────────
class CrystalSDKBridge:
    _COM_PROGIDS = [
        "CrystalDecisions.CrystalReports.Engine.ReportDocument.2020.0",
        "CrystalDecisions.CrystalReports.Engine.ReportDocument.14.0",
        "CrystalDecisions.CrystalReports.Engine.ReportDocument.13.0",
        "CrystalDecisions.CrystalReports.Engine.ReportDocument",
    ]
    _NET_ASSEMBLIES = [
        "CrystalDecisions.CrystalReports.Engine",
        "CrystalDecisions.Shared",
    ]

    def load(self, path: str, r: CrystalReport):
        if not CRYSTAL_SDK_AVAILABLE:
            raise RuntimeError("No Crystal SDK found (pythonnet and win32com both absent)")
        if CRYSTAL_SDK_MODE == "clr":
            self._load_clr(path, r)
        else:
            self._load_com(path, r)

    # ── pythonnet path ───────────────────────────────────────────────────────
    def _load_clr(self, path: str, r: CrystalReport):
        import clr, sys as _sys

        override = os.environ.get("CRYSTAL_ASSEMBLY_PATH", "").strip()
        dirs_to_try = [override] if override else self._find_all_asm_dirs()

        # Filter to dirs that actually contain the Engine DLL
        engine_dirs = [
            d for d in dirs_to_try
            if os.path.isdir(d) and
               os.path.isfile(os.path.join(d, "CrystalDecisions.CrystalReports.Engine.dll"))
        ]

        loaded = False
        last_err = None

        for d in engine_dirs:
            try:
                # Add dir (and registry CommonFiles dir) to sys.path so the
                # .NET runtime can resolve peer DLLs at bind time
                for extra in [d] + self._find_extra_dep_dirs(d):
                    if extra and extra not in _sys.path:
                        _sys.path.insert(0, extra)

                # Load Engine by full absolute path — bypasses GAC version conflicts
                engine_dll = os.path.join(d, "CrystalDecisions.CrystalReports.Engine.dll")
                clr.AddReference(engine_dll)
                emit(f"  CLR: loaded Engine from {d}")

                # Eagerly load every Crystal DLL in the same dir so the .NET
                # runtime doesn't have to resolve them lazily from a different location
                for dll_name in [
                    "CrystalDecisions.Shared.dll",
                    "CrystalDecisions.ReportAppServer.CommLayer.dll",
                    "CrystalDecisions.ReportAppServer.Controllers.dll",
                    "CrystalDecisions.ReportAppServer.ClientDoc.dll",
                    "CrystalDecisions.ReportAppServer.DataSetConversion.dll",
                    "CrystalDecisions.Windows.Forms.dll",
                    "CrystalDecisions.Web.dll",
                ]:
                    full = os.path.join(d, dll_name)
                    if os.path.isfile(full):
                        try:
                            clr.AddReference(full)
                        except Exception:
                            pass  # optional — swallow silently

                loaded = True
                break

            except Exception as e:
                last_err = e
                emit(f"  CLR: failed for {d}: {e}", "WARN")

        # Pure GAC fallback — try by strong name (no dir needed)
        if not loaded:
            try:
                clr.AddReference("CrystalDecisions.CrystalReports.Engine")
                emit("  CLR: loaded from GAC by strong name")
                loaded = True
            except Exception as e:
                last_err = e
                emit(f"  CLR: GAC strong-name load failed: {e}", "WARN")

        if not loaded:
            raise RuntimeError(
                f"Could not load Crystal .NET assemblies (last: {last_err}).\n"
                "Run GET /diagnose for a scan report, then set env var:\n"
                "  set CRYSTAL_ASSEMBLY_PATH=<folder with "
                "CrystalDecisions.CrystalReports.Engine.dll>"
            )

        try:
            from CrystalDecisions.CrystalReports.Engine import ReportDocument  # type: ignore
        except ImportError as e:
            raise RuntimeError(
                f"Assembly loaded but Python import still failed: {e}\n"
                "A dependency DLL is likely missing.  Check CRYSTAL_ASSEMBLY_PATH "
                "or run GET /diagnose."
            )

        rd = ReportDocument()
        rd.Load(path)
        emit(f"  CLR: opened {Path(path).name}")
        self._extract(rd, r, path)
        rd.Close()

    @staticmethod
    def _find_extra_dep_dirs(engine_dir: str) -> List[str]:
        """Return additional dirs from the registry that hold Crystal dependency DLLs."""
        extras: List[str] = []
        try:
            import winreg
            with winreg.OpenKey(
                    winreg.HKEY_LOCAL_MACHINE,
                    r"SOFTWARE\SAP BusinessObjects\Suite XI 4.0\Crystal Reports") as k:
                for vn in ("CommonFiles", "Path", "ChartSupportPath"):
                    try:
                        val, _ = winreg.QueryValueEx(k, vn)
                        d = str(val).strip().rstrip("\\")
                        if d and os.path.isdir(d) and d != engine_dir:
                            extras.append(d)
                    except OSError:
                        pass
        except Exception:
            pass
        return extras

    @staticmethod
    def _find_all_asm_dirs() -> List[str]:
        """
        Return every candidate dir in priority order:
          1. Registry-derived paths (accurate for the actual install)
          2. Program Files filesystem scan
          3. GAC subdirectories
          4. Script dir / CWD
        """
        dirs: List[str] = []

        if platform.system() == "Windows":
            try:
                import winreg

                reg_searches = [
                    # (hive, key, [value names to probe])
                    (winreg.HKEY_LOCAL_MACHINE,
                     r"SOFTWARE\SAP BusinessObjects\Suite XI 4.0\Crystal Reports",
                     ["Path", "CommonFiles", "ChartSupportPath"]),
                    (winreg.HKEY_LOCAL_MACHINE,
                     r"SOFTWARE\SAP BusinessObjects\Crystal Reports for Visual Studio\Default",
                     ["", "InstallDir", "Path"]),
                    (winreg.HKEY_LOCAL_MACHINE,
                     r"SOFTWARE\SAP BusinessObjects\Crystal Reports for Visual Studio",
                     ["", "InstallDir", "Path"]),
                    (winreg.HKEY_LOCAL_MACHINE,
                     r"SOFTWARE\WOW6432Node\SAP BusinessObjects\Crystal Reports for Visual Studio\Default",
                     ["", "InstallDir", "Path"]),
                    (winreg.HKEY_LOCAL_MACHINE,
                     r"SOFTWARE\WOW6432Node\SAP BusinessObjects\Suite XI 4.0\Crystal Reports",
                     ["Path", "CommonFiles"]),
                    (winreg.HKEY_LOCAL_MACHINE,
                     r"SOFTWARE\Business Objects\Crystal Reports for Visual Studio\Default",
                     ["", "InstallDir", "Path"]),
                    (winreg.HKEY_CURRENT_USER,
                     r"SOFTWARE\SAP BusinessObjects\Suite XI 4.0\Crystal Reports",
                     ["Path", "CommonFiles"]),
                ]

                for hive, rk, val_names in reg_searches:
                    try:
                        with winreg.OpenKey(hive, rk) as k:
                            for vn in val_names:
                                try:
                                    val, _ = winreg.QueryValueEx(k, vn)
                                    base = str(val).strip().rstrip("\\")
                                    if not base:
                                        continue
                                    for sub in ["", "win32_x86", "win64_x64", "bin", "x64"]:
                                        d = os.path.join(base, sub) if sub else base
                                        if os.path.isdir(d):
                                            dirs.append(d)
                                except OSError:
                                    pass
                    except (FileNotFoundError, OSError):
                        pass

            except ImportError:
                pass

        # Filesystem scan
        import glob
        roots = [
            os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"),
            os.environ.get("ProgramFiles",      r"C:\Program Files"),
            r"C:\Program Files (x86)",
            r"C:\Program Files",
        ]
        patterns = [
            r"SAP BusinessObjects\Crystal Reports*",
            r"SAP BusinessObjects\SAP BusinessObjects Enterprise XI*\win32_x86",
            r"SAP BusinessObjects\SAP BusinessObjects Enterprise XI*\win64_x64",
            r"Business Objects\Crystal Reports*",
            r"Crystal Reports*",
        ]
        for root in (r for r in roots if r and os.path.isdir(r)):
            for pat in patterns:
                for hit in glob.glob(os.path.join(root, pat)):
                    if os.path.isdir(hit):
                        dirs.append(hit)
                    for sub in glob.glob(os.path.join(hit, "*")):
                        if os.path.isdir(sub):
                            dirs.append(sub)

        # GAC
        gac = r"C:\Windows\Microsoft.NET\assembly\GAC_MSIL"
        if os.path.isdir(gac):
            for e in glob.glob(os.path.join(gac, "CrystalDecisions*", "*")):
                if os.path.isdir(e):
                    dirs.append(e)

        dirs.append(os.path.dirname(os.path.abspath(__file__)))
        dirs.append(os.getcwd())

        seen: set = set()
        result = []
        for d in dirs:
            nd = os.path.normpath(d)
            if nd not in seen:
                seen.add(nd); result.append(nd)
        return result

    # ── win32com path ────────────────────────────────────────────────────────
    def _load_com(self, path: str, r: CrystalReport):
        import win32com.client as win32
        rd, last_err = None, None
        for progid in self._COM_PROGIDS:
            try:
                rd = win32.Dispatch(progid)
                emit(f"  COM ProgID: {progid}")
                break
            except Exception as e:
                last_err = e
        if rd is None:
            raise RuntimeError(
                f"COM dispatch failed for all ProgIDs (last: {last_err}).\n"
                "The .NET SDK does not register COM ProgIDs.\n"
                "Fix: pip install pythonnet  — or place a reference .pbit "
                "alongside the .rpt to seed the schema without the SDK."
            )
        rd.Load(path)
        self._extract(rd, r, path)
        rd.Close()

    # ── shared extraction ────────────────────────────────────────────────────
    def _extract(self, rd, r: CrystalReport, path: str):
        """
        Extract all report design elements from the loaded ReportDocument.
        Every attribute access is individually guarded — property availability
        varies between SDK v13 and v14 and between COM and CLR modes.
        """

        def _s(obj, *attrs, default=""):
            """Safely read the first resolvable dot-path attribute as string."""
            for a in attrs:
                try:
                    v = obj
                    for part in a.split("."):
                        v = getattr(v, part)
                    return str(v) if v is not None else default
                except Exception:
                    pass
            return default

        def _i(obj, attr, default=0):
            try:    return int(getattr(obj, attr))
            except: return default

        def _b(obj, attr, default=False):
            try:    return bool(getattr(obj, attr))
            except: return default

        def _color(val) -> str:
            """Convert Crystal color int to #RRGGBB hex."""
            try:
                n = int(val)
                if n < 0: n = n & 0xFFFFFF
                return f"#{n:06X}"
            except:
                return ""

        # ── Summary info ──────────────────────────────────────────────────────
        try:
            si = rd.SummaryInfo
            r.report_title = _s(si, "ReportTitle", "Title", default=Path(path).stem)
            r.subject       = _s(si, "Subject", "Comments", "Description")
            r.author        = _s(si, "Author", "CreatedBy")
        except Exception:
            r.report_title = Path(path).stem
        emit(f"  CLR: title={r.report_title!r}")

        # ── Database tables + connection info + SQL command text ──────────────
        try:
            db = rd.Database
            for i in range(db.Tables.Count):
                try:
                    t  = db.Tables[i]
                    li = t.LogOnInfo
                    ci = li.ConnectionInfo

                    srv  = _s(ci, "ServerName",   "Server",   default="")
                    dbn  = _s(ci, "DatabaseName", "Database", default="")
                    ctype= _s(ci, "DatabaseDLL", "Type", default="")
                    conn_str = self._conn(li)

                    # When DatabaseName is empty (CRQE/ODBC connections), resolve
                    # the physical .mdb path through multiple fallback strategies.
                    # Note: dbn starts empty here.
                    if not dbn:
                        # Strategy 1: Windows registry — read DBQ from ODBC DSN definition.
                        # This is the most reliable source since Crystal registers the DSN
                        # with the actual file path.
                        if srv and platform.system() == "Windows":
                            reg_path = self._dsn_to_mdb_path(srv)
                            if reg_path:
                                dbn = reg_path
                                emit(f"  CLR: table[{i}] path from ODBC registry DSN "
                                     f"'{srv}'={dbn!r}")

                        # Strategy 2: t.Location SDK property — only accept if it looks
                        # like a file path (contains a path separator or file extension)
                        if not dbn:
                            try:
                                loc = _s(t, "Location", default="")
                                # Accept only if it looks like a real path:
                                # must contain a slash/backslash OR a file extension,
                                # and must not be a bare table name or "(unknown)"
                                import re as _re
                                is_path = (
                                    loc
                                    and loc.lower() not in ("(unknown)", "unknown", "")
                                    and not loc.lstrip().upper().startswith("SELECT")
                                    and (
                                        any(sep in loc for sep in (r"\\", "/", ":\\"))
                                        or _re.search(r'\.\w{2,5}$', loc)
                                    )
                                )
                                if is_path:
                                    dbn = loc
                                    emit(f"  CLR: table[{i}] path from Location={dbn!r}")
                                elif loc:
                                    emit(f"  CLR: table[{i}] Location={loc!r} rejected "
                                         f"(not a file path — looks like table name)")
                            except Exception:
                                pass

                    # Log exactly what the SDK gave us — helps debug M query issues
                    emit(f"  CLR: table[{i}] name={_s(t,'Name')!r}  "
                         f"ServerName={srv!r}  DatabaseName={dbn!r}  "
                         f"DatabaseDLL={ctype!r}")

                    # Build connection string from resolved values
                    conn_str = self._conn_from_parts(srv, dbn)
                    ct = CrystalTable(
                        name=_s(t, "Name", default=f"Table{i}"),
                        alias=_s(t, "Alias", "Name", default=f"Table{i}"),
                        db_type=self._dbtype(dbn, srv, ctype),
                        connection_string=conn_str,
                        server=srv,
                        database=dbn,
                        connection_type=ctype,
                    )

                    # SQL command text (CommandTable)
                    try:
                        cmd = _s(t, "CommandText", "SQLExpression", default="")
                        if not cmd:
                            # Some versions expose it via Location
                            loc = _s(t, "Location", default="")
                            if loc.lstrip().upper().startswith("SELECT"):
                                cmd = loc
                        ct.sql_command = cmd
                    except Exception:
                        pass

                    # Fields belonging to this table — try three strategies.
                    # Track seen names so strategies don't double-add the same field.
                    tname = str(t.Name)
                    seen_field_names: set = set()

                    def _add_field(f_obj):
                        """Add a SDK field object to ct + r if not already seen."""
                        try:
                            fname = str(f_obj.Name).strip()
                            if not fname or fname in seen_field_names:
                                return False
                            # Resolve data type — FieldValueType may throw on some CLR builds
                            try:
                                dtype = self._ft(f_obj.FieldValueType)
                            except Exception:
                                dtype = "string"
                            cf = CrystalField(
                                name=fname,
                                field_type="database",
                                data_type=dtype,
                                table_name=tname,
                                source_column=fname,
                            )
                            # Only mark seen + append after successful construction
                            seen_field_names.add(fname)
                            ct.fields.append(cf)
                            r.fields.append(cf)
                            return True
                        except Exception:
                            return False

                    # Strategy 1: table.Fields (direct — most reliable in CLR v13/v14)
                    try:
                        for j in range(t.Fields.Count):
                            try: _add_field(t.Fields[j])
                            except Exception: pass
                    except Exception:
                        pass

                    # Strategy 2: DataDefinition.DatabaseFields filtered by table name
                    if not seen_field_names:
                        try:
                            for j in range(rd.DataDefinition.DatabaseFields.Count):
                                try:
                                    f = rd.DataDefinition.DatabaseFields[j]
                                    ft = _s(f, "TableName", default="").strip()
                                    if ft.lower() in (tname.lower(), ""):
                                        _add_field(f)
                                except Exception:
                                    pass
                        except Exception:
                            pass

                    # Strategy 3: DataDefinition.Fields (all field types) filtered by table name
                    if not seen_field_names:
                        try:
                            for j in range(rd.DataDefinition.Fields.Count):
                                try:
                                    f = rd.DataDefinition.Fields[j]
                                    ft = _s(f, "TableName", default="").strip()
                                    if ft.lower() in (tname.lower(), ""):
                                        _add_field(f)
                                except Exception:
                                    pass
                        except Exception:
                            pass

                    n_found = len(seen_field_names)
                    if n_found == 0:
                        emit(f"  CLR: table '{tname}' — no fields (all 3 strategies failed)", "WARN")
                    else:
                        emit(f"  CLR: table '{tname}' → {n_found} field(s)")

                    r.tables.append(ct)
                except Exception as e:
                    emit(f"  CLR: table[{i}] skipped: {e}", "WARN")
        except Exception as e:
            emit(f"  CLR: Database unavailable: {e}", "WARN")

        # ── Formula fields (with full formula code) ───────────────────────────
        try:
            ff_count = rd.DataDefinition.FormulaFields.Count
            emit(f"  CLR: FormulaFields.Count={ff_count}")
            for i in range(ff_count):
                try:
                    ff  = rd.DataDefinition.FormulaFields[i]
                    txt = _s(ff, "Text", "FormulaText", "Formula")
                    raw_name = str(ff.Name)
                    clean_name = raw_name.lstrip("{@").rstrip("}")
                    # FormulaFieldDefinition uses ValueType, not FieldValueType
                    try:
                        dtype = self._ft(ff.ValueType)
                    except AttributeError:
                        try:
                            dtype = self._ft(ff.FieldValueType)
                        except AttributeError:
                            dtype = "string"
                    r.fields.append(CrystalField(
                        name=clean_name,
                        field_type="formula",
                        data_type=dtype,
                        formula_text=txt,
                    ))
                    emit(f"  CLR: formula[{i}] name={clean_name!r} text={txt[:60]!r}")
                except Exception as e:
                    emit(f"  CLR: formula[{i}] skipped: {e}", "WARN")
        except Exception as e:
            emit(f"  CLR: FormulaFields unavailable: {e}", "WARN")

        # ── Parameter definitions ─────────────────────────────────────────────
        try:
            pf_count = rd.DataDefinition.ParameterFields.Count
            emit(f"  CLR: ParameterFields.Count={pf_count}")
            for i in range(pf_count):
                try:
                    pf = rd.DataDefinition.ParameterFields[i]
                    defaults = []
                    try:
                        for d in range(pf.DefaultValues.Count):
                            defaults.append(str(pf.DefaultValues[d]))
                    except Exception:
                        pass
                    # ParameterValueType property name varies by SDK version
                    try:
                        ptype = self._ft(pf.ParameterValueType)
                    except AttributeError:
                        try:
                            ptype = self._ft(pf.ValueType)
                        except AttributeError:
                            ptype = "string"
                    param = {
                        "name":           _s(pf, "Name"),
                        "type":           ptype,
                        "prompt":         _s(pf, "PromptText", "Name"),
                        "allow_multiple": _b(pf, "AllowMultipleValues"),
                        "allow_range":    _b(pf, "AllowRangeValue"),
                        "defaults":       defaults,
                    }
                    r.parameters.append(param)
                    emit(f"  CLR: param[{i}] name={param['name']!r} type={param['type']!r}")
                except Exception as e:
                    emit(f"  CLR: param[{i}] skipped: {e}", "WARN")
        except Exception as e:
            emit(f"  CLR: ParameterFields unavailable: {e}", "WARN")

        # ── Record + group selection formulas ─────────────────────────────────
        try:
            r.record_selection_formula = str(
                rd.DataDefinition.RecordSelectionFormula or "")
            if r.record_selection_formula:
                emit(f"  CLR: RecordSelectionFormula={r.record_selection_formula[:60]!r}")
        except Exception as e:
            emit(f"  CLR: RecordSelectionFormula unavailable: {e}", "WARN")
        try:
            r.group_selection_formula = str(
                rd.DataDefinition.GroupSelectionFormula or "")
        except Exception:
            pass

        # ── Grouping and sorting ──────────────────────────────────────────────
        try:
            g_count = rd.DataDefinition.Groups.Count
            emit(f"  CLR: Groups.Count={g_count}")
            for i in range(g_count):
                try:
                    g     = rd.DataDefinition.Groups[i]
                    # ConditionField property name varies — try multiple paths
                    fname = _s(g, "ConditionField.Name", "ConditionField.FieldName",
                               "ConditionField", default="")
                    if not fname:
                        try:
                            cf = g.ConditionField
                            fname = str(cf.Name) if hasattr(cf,'Name') else str(cf)
                        except Exception:
                            fname = f"Group{i}"
                    try:
                        sd_raw = g.SortDirection if hasattr(g, "SortDirection") \
                                 else getattr(g, "ConditionField.SortDirection", 0)
                        order = "desc" if int(sd_raw) in (1, 3) else "asc"
                    except Exception:
                        order = "asc"
                    r.groups.append({
                        "field":       fname,
                        "order":       order,
                        "group_level": i,
                    })
                    emit(f"  CLR: group[{i}] field={fname!r} order={order!r}")
                except Exception as e:
                    emit(f"  CLR: group[{i}] skipped: {e}", "WARN")
        except Exception as e:
            emit(f"  CLR: Groups unavailable: {e}", "WARN")

        # Sort fields (record-level sorting)
        try:
            sf_count = rd.DataDefinition.SortFields.Count
            emit(f"  CLR: SortFields.Count={sf_count}")
            for i in range(sf_count):
                try:
                    sf  = rd.DataDefinition.SortFields[i]
                    # Field property name varies across SDK versions
                    fn = _s(sf, "Field.Name", "Field.FieldName", "FieldName", default="")
                    if not fn:
                        try:
                            fld = sf.Field
                            fn = str(fld.Name) if hasattr(fld,'Name') else str(fld)
                        except Exception:
                            fn = f"SortField{i}"
                    try:
                        sd  = int(sf.SortDirection)
                        direction = "desc" if sd == 1 else "asc"
                    except Exception:
                        direction = "asc"
                    try:
                        st = _s(sf, "SortType", default="record").lower()
                    except Exception:
                        st = "record"
                    r.sort_fields.append(CrystalSortField(fn, direction, st))
                    emit(f"  CLR: sort[{i}] field={fn!r} dir={direction!r}")
                except Exception as e:
                    emit(f"  CLR: sort[{i}] skipped: {e}", "WARN")
        except Exception as e:
            emit(f"  CLR: SortFields unavailable: {e}", "WARN")

        # ── Sections, objects, and subreports ─────────────────────────────────
        try:
            for i in range(rd.ReportDefinition.Sections.Count):
                try:
                    s   = rd.ReportDefinition.Sections[i]
                    cs  = CrystalSection(
                        name=_s(s, "Name", default=f"Section{i}"),
                        section_type=self._st(s.Kind),
                    )
                    try:
                        fmt = s.SectionFormat
                        cs.suppress         = _b(fmt, "EnableSuppress")
                        cs.suppress_formula = _s(fmt, "EnableSuppressFormula",
                                                  "SuppressFormula", default="")
                        cs.height           = _i(fmt, "Height")
                        try:
                            cs.background_color = _color(fmt.BackgroundColor)
                        except Exception:
                            pass
                    except Exception:
                        pass

                    # Report objects in this section
                    try:
                        for j in range(s.ReportObjects.Count):
                            od = self._obj(s.ReportObjects[j])
                            if od:
                                cs.objects.append(od)
                                # Track subreports
                                if od.get("type") == "subreport":
                                    sr = CrystalSubreport(
                                        name=od.get("name", ""),
                                        section=cs.section_type,
                                    )
                                    sr.x = od.get("left", 0)
                                    sr.y = od.get("top",  0)
                                    sr.w = od.get("width",  0)
                                    sr.h = od.get("height", 0)
                                    # Extract link fields
                                    try:
                                        sub_rd = rd.OpenSubreport(od["name"])
                                        lf_col = sub_rd.DataDefinition.LinkFields
                                        for lk in range(lf_col.Count):
                                            try:
                                                link = lf_col[lk]
                                                sr.link_fields.append({
                                                    "main_field": _s(link, "MainReportFieldName", default=""),
                                                    "sub_field":  _s(link, "SubreportFieldName",  default=""),
                                                })
                                            except Exception:
                                                pass
                                    except Exception:
                                        pass
                                    r.subreports.append(sr)
                    except Exception:
                        pass

                    r.sections.append(cs)
                except Exception as e:
                    emit(f"  CLR: section[{i}] skipped: {e}", "WARN")
        except Exception as e:
            emit(f"  CLR: Sections unavailable: {e}", "WARN")

        db_col_count = sum(len(t.fields) for t in r.tables)
        formula_count = sum(1 for f in r.fields if f.field_type == "formula")
        emit(f"  CLR: {len(r.tables)} table(s)  "
             f"{db_col_count} DB cols  "
             f"{formula_count} formulas  "
             f"{len(r.groups)} group(s)  {len(r.sort_fields)} sort(s)  "
             f"{len(r.subreports)} subreport(s)  "
             f"{len(r.sections)} section(s)")

    @staticmethod
    def _dsn_to_mdb_path(dsn_name: str) -> str:
        """
        Look up the physical .mdb/.accdb file path for an ODBC DSN from the
        Windows registry.  Crystal Reports ODBC connections store the DSN name
        in ServerName but leave DatabaseName empty; the actual path is in the
        DSN definition under HKLM or HKCU ODBC.INI.

        Returns the path string, or "" if not found.
        """
        if platform.system() != "Windows":
            return ""
        try:
            import winreg
            hive_paths = [
                (winreg.HKEY_LOCAL_MACHINE,
                 rf"SOFTWARE\ODBC\ODBC.INI\{dsn_name}",
                 "HKLM"),
                (winreg.HKEY_LOCAL_MACHINE,
                 rf"SOFTWARE\WOW6432Node\ODBC\ODBC.INI\{dsn_name}",
                 "HKLM WOW64"),
                (winreg.HKEY_CURRENT_USER,
                 rf"SOFTWARE\ODBC\ODBC.INI\{dsn_name}",
                 "HKCU"),
            ]
            for hive, path, hive_label in hive_paths:
                try:
                    with winreg.OpenKey(hive, path) as k:
                        emit(f"  Registry: found DSN '{dsn_name}' under {hive_label}")
                        # Enumerate all values to log them for debugging
                        vals = {}
                        idx = 0
                        while True:
                            try:
                                n, v, _ = winreg.EnumValue(k, idx)
                                vals[n] = str(v)
                                idx += 1
                            except OSError:
                                break
                        emit(f"  Registry: DSN values = {vals}")
                        # Access ODBC driver stores path in "DBQ" value
                        for val_name in ("DBQ", "Database", "Dbq", "dbq"):
                            if val_name in vals:
                                v = vals[val_name].strip()
                                if v and v.lower() not in ("(unknown)", "unknown", ""):
                                    return v
                except OSError:
                    pass
            emit(f"  Registry: DSN '{dsn_name}' not found in HKLM or HKCU ODBC.INI")
        except Exception as e:
            emit(f"  Registry: lookup failed: {e}", "WARN")
        return ""

    @staticmethod
    def _conn(li) -> str:
        ci = li.ConnectionInfo
        parts = []
        if ci.ServerName:   parts.append(f"server={ci.ServerName}")
        if ci.DatabaseName: parts.append(f"database={ci.DatabaseName}")
        return ";".join(parts)

    @staticmethod
    def _conn_from_parts(srv: str, dbn: str) -> str:
        parts = []
        if srv: parts.append(f"server={srv}")
        if dbn: parts.append(f"database={dbn}")
        return ";".join(parts)

    @staticmethod
    def _dbtype(dbn: str, srv: str = "", dll: str = "") -> str:
        """
        Determine db_type from SDK connection info.
        dbn = DatabaseName (often the .mdb path for Access)
        srv = ServerName   (DSN name for ODBC, server for SQL)
        dll = DatabaseDLL  (Crystal's driver code: CRQE=ODBC, SQLSRV=SQL Server, etc.)
        """
        combined = (dbn + srv).lower()
        if ".mdb" in combined or ".accdb" in combined: return "access"
        if ".xls" in combined:                         return "excel"
        # CRQE = Crystal Reports Query Engine (ODBC driver)
        # If DatabaseDLL contains CRQE or ODBC, treat as ODBC with DSN in ServerName
        dll_low = dll.lower()
        if "crqe" in dll_low or "odbc" in dll_low:    return "odbc"
        # If DatabaseName is empty but ServerName is set → likely ODBC DSN
        if not dbn and srv:                            return "odbc"
        if not dbn and not srv:                        return "unknown"
        return "sql"

    @staticmethod
    def _ft(vt) -> str:
        # Handle both int enum values (COM) and .NET enum objects (clr)
        try:
            return {0:"string",1:"number",2:"currency",3:"boolean",
                    4:"date",5:"time",6:"datetime"}.get(int(vt), "string")
        except (TypeError, ValueError):
            name = str(vt).lower()
            if "currency" in name:               return "currency"
            if "number" in name or "int" in name: return "number"
            if "boolean" in name:                return "boolean"
            if "datetime" in name:               return "datetime"
            if "date" in name:                   return "date"
            if "time" in name:                   return "datetime"
            return "string"

    @staticmethod
    def _st(k) -> str:
        try: k = int(k)
        except: k = -1
        return {0:"reportHeader",1:"pageHeader",2:"groupHeader",3:"detail",
                4:"groupFooter",5:"pageFooter",6:"reportFooter"}.get(k, "detail")

    @staticmethod
    def _obj(obj) -> Optional[dict]:
        """
        Extract a report object into a plain dict.
        Kind codes: 1=Field, 2=Text, 3=Line, 4=Box, 5=Chart, 6=CrossTab,
                    7=Subreport, 8=OLE, 9=Map, 10=FieldHeading
        """
        try:
            k = int(obj.Kind)
            b: dict = {
                "kind":   k,
                "name":   str(obj.Name),
                "left":   int(obj.Left),
                "top":    int(obj.Top),
                "width":  int(obj.Width),
                "height": int(obj.Height),
            }

            if k == 1:      # Database / formula field
                b["type"] = "field"
                try: b["field_name"] = str(obj.DataSource.Name)
                except Exception: b["field_name"] = ""
                try: b["field_type"] = str(obj.DataSource.Kind)
                except Exception: pass
                # Format properties
                try:
                    fmt = obj.FieldFormat
                    b["format_string"] = str(fmt.NumericFormat.CustomFormatString
                                             if hasattr(fmt, "NumericFormat") else "")
                except Exception: pass
                # Font
                try:
                    fb = obj.FontColor
                    b["font_color"] = f"#{int(fb) & 0xFFFFFF:06X}"
                except Exception: pass

            elif k == 2:    # Text object
                b["type"] = "text"
                try: b["text"] = str(obj.Text)
                except Exception: b["text"] = ""
                try:
                    b["font_name"]  = str(obj.Font.Name)
                    b["font_size"]  = int(obj.Font.Size)
                    b["bold"]       = bool(obj.Font.Bold)
                    b["italic"]     = bool(obj.Font.Italic)
                except Exception: pass

            elif k == 3:    # Line
                b["type"] = "line"
                try: b["line_style"] = int(obj.LineStyle)
                except Exception: pass
                try: b["line_width"] = int(obj.LineThickness)
                except Exception: pass

            elif k == 4:    # Box / Rectangle
                b["type"] = "box"
                try: b["line_style"] = int(obj.LineStyle)
                except Exception: pass
                try:
                    bc = obj.FillColor
                    b["fill_color"] = f"#{int(bc) & 0xFFFFFF:06X}"
                except Exception: pass

            elif k == 5:    # Chart
                b["type"] = "chart"
                try: b["chart_type"] = str(obj.Chart.Type)
                except Exception: b["chart_type"] = "unknown"
                try: b["chart_title"] = str(obj.Chart.Title)
                except Exception: pass

            elif k == 6:    # Cross-tab
                b["type"] = "crosstab"

            elif k == 7:    # Subreport
                b["type"] = "subreport"
                try: b["name"] = str(obj.SubreportName)
                except Exception: pass
                try: b["on_demand"] = bool(obj.IsOnDemand)
                except Exception: pass

            elif k == 10:   # Field heading / column header text
                b["type"] = "field_heading"
                try: b["text"] = str(obj.Text)
                except Exception: b["text"] = ""

            else:
                b["type"] = f"kind_{k}"

            return b
        except Exception:
            return None

# ── Formula Translator ───────────────────────────────────────────────────────
class FormulaTranslator:
    _FN={"ToText":"TEXT","ToNumber":"VALUE","StrLen":"LEN","Mid":"MID","Left":"LEFT","Right":"RIGHT",
         "Trim":"TRIM","Upper":"UPPER","Lower":"LOWER","InStr":"FIND","CDate":"DATEVALUE","Year":"YEAR",
         "Month":"MONTH","Day":"DAY","DateAdd":"DATEADD","DateDiff":"DATEDIFF","Now":"NOW","Today":"TODAY",
         "IsNull":"ISBLANK","Round":"ROUND","Abs":"ABS","Int":"INT","Mod":"MOD","Chr":"UNICHAR","Asc":"UNICODE"}
    _AG={"Sum":"SUM","Average":"AVERAGE","Count":"COUNT","Maximum":"MAX","Minimum":"MIN","StdDev":"STDEV.P","Variance":"VAR.P"}
    def to_dax(self,formula,table="Table"):
        if not formula: return ""
        d=formula.strip(); d=re.sub(r"^formula\s*=\s*","",d,flags=re.IGNORECASE)
        d=re.sub(r"\{(\w+)\.(\w+)\}",lambda m:f"'{m.group(1)}'[{m.group(2)}]",d)
        d=re.sub(r"\{(\w+)\}",lambda m:f"'{table}'[{m.group(1)}]",d); d=self._ite(d)
        for cf,df in self._FN.items(): d=re.sub(rf"\b{re.escape(cf)}\s*\(",f"{df}(",d,flags=re.IGNORECASE)
        for ca,da in self._AG.items():
            d=re.sub(rf"\b{ca}\s*\(\s*\{{(\w+)\.(\w+)\}}\s*\)",lambda m,a=da:f"{a}('{m.group(1)}'[{m.group(2)}])",d,flags=re.IGNORECASE)
        d=re.sub(r"\band\b","&&",d,flags=re.IGNORECASE); d=re.sub(r"\bor\b","||",d,flags=re.IGNORECASE)
        return d.strip()
    def to_m(self,formula):
        if not formula: return "// No filter"
        m=formula.strip()
        m=re.sub(r"\{(\w+)\.(\w+)\}",r"[\2]",m); m=re.sub(r"\{(\w+)\}",r"[\1]",m)
        m=re.sub(r"\bIf\b","if",m,flags=re.IGNORECASE); m=re.sub(r"\bThen\b","then",m,flags=re.IGNORECASE)
        m=re.sub(r"\bElse\b","else",m,flags=re.IGNORECASE)
        m=re.sub(r"\bIsNull\s*\(([^)]+)\)",r"\1 = null",m,flags=re.IGNORECASE)
        return f"Table.SelectRows(Source, each\n    {m}\n)"

    def to_m_filter(self, formula: str) -> str:
        """
        Translate a Crystal record-selection formula to an M filter predicate
        suitable for use inside Table.SelectRows(tbl, each <predicate>).
        Returns empty string if the formula cannot be translated.
        """
        if not formula or not formula.strip():
            return ""
        m = formula.strip()
        # Remove outer braces if present
        m = re.sub(r"^formula\s*=\s*", "", m, flags=re.IGNORECASE)
        # {Table.Field} → [Field]
        m = re.sub(r"\{(\w+)\.(\w+)\}", r"[\2]", m)
        # {Field} → [Field]
        m = re.sub(r"\{(\w+)\}", r"[\1]", m)
        # Crystal boolean operators
        m = re.sub(r"\band\b", "and", m, flags=re.IGNORECASE)
        m = re.sub(r"\bor\b",  "or",  m, flags=re.IGNORECASE)
        m = re.sub(r"\bnot\b", "not", m, flags=re.IGNORECASE)
        # IsNull → = null
        m = re.sub(r"\bIsNull\s*\(([^)]+)\)", r"\1 = null", m, flags=re.IGNORECASE)
        # Crystal equality uses = not ==
        return m.strip()
    @staticmethod
    def _ite(expr):
        p=re.compile(r"\bIf\b\s+(.+?)\s+\bThen\b\s+(.+?)(?:\s+\bElse\b\s+(.+))?$",re.IGNORECASE|re.DOTALL)
        m=p.match(expr.strip())
        if m:
            e_=m.group(3).strip() if m.group(3) else "BLANK()"
            return f"IF({m.group(1).strip()}, {m.group(2).strip()}, {e_})"
        return expr


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT ANALYZER — decides PBIT vs RDL based on report characteristics
# ═══════════════════════════════════════════════════════════════════════════════
class ReportAnalyzer:
    """
    Scores a CrystalReport on features that indicate pixel-perfect / paginated
    layout (→ RDL) vs interactive analytics (→ PBIT).

    Scoring rubric (positive = RDL, negative = PBIT):
      +20  Has subreports (RDL handles them natively as drill-through)
      +15  Has cross-tab objects
      +10  Has page header or page footer with content
      +10  Report has 3+ group levels (complex grouping favors Matrix / RDL)
      +8   Has lines or boxes (pixel-perfect formatting indicators)
      +5   Has record selection formula (complex filter → prefer M in PBIT,
               but strongly formatted reports still go RDL)
      +5   Has chart objects (charts are fine in both, mild RDL preference)
      +3   Per subreport beyond the first
      +3   Per group level beyond 2
      -15  Has only a detail section and no groups (simple tabular → PBIT)
      -10  Has fewer than 3 groups and no subreports (flat list → PBIT)
      -8   Formula fields that are aggregations (SUM/COUNT → DAX Measures, PBIT)
      -5   Single datasource, no SQL command (simple query → PBIT)

    Threshold: score >= 20 → RDL;  score < 20 → PBIT
    """

    RDL_THRESHOLD = 20

    def analyze(self, r: CrystalReport) -> str:
        """Set r.recommended_output, r.routing_reason, r.routing_scores. Returns "pbit"|"rdl"."""
        scores: dict = {}

        # Subreports
        sr_count = len(r.subreports)
        if sr_count > 0:
            scores["subreports"]       = 20
            scores["extra_subreports"] = (sr_count - 1) * 3

        # Cross-tabs
        xtab = sum(1 for sec in r.sections
                   for obj in sec.objects if obj.get("type") == "crosstab")
        if xtab:
            scores["crosstab"] = 15

        # Page header / footer with content
        pg_sections = [s for s in r.sections
                       if s.section_type in ("pageHeader", "pageFooter")
                       and len(s.objects) > 0]
        if pg_sections:
            scores["page_header_footer"] = 10

        # Group levels
        g_count = len(r.groups)
        if g_count >= 3:
            scores["many_groups"] = 10 + (g_count - 2) * 3
        elif g_count == 2:
            scores["two_groups"] = 5

        # Lines / boxes (layout precision indicator)
        layout_objs = sum(1 for sec in r.sections
                          for obj in sec.objects
                          if obj.get("type") in ("line", "box"))
        if layout_objs > 0:
            scores["layout_objects"] = min(layout_objs * 2, 8)

        # Charts
        charts = sum(1 for sec in r.sections
                     for obj in sec.objects if obj.get("type") == "chart")
        if charts:
            scores["charts"] = 5

        # Record selection formula
        if r.record_selection_formula:
            scores["record_selection"] = 5

        # Flat / simple report indicators (negative scores → PBIT)
        has_groups = g_count > 0
        has_detail_only = (not has_groups and not sr_count and not xtab)
        if has_detail_only:
            scores["flat_detail_only"] = -15
        elif not has_groups and not sr_count:
            scores["no_groups_no_subreports"] = -10

        # Aggregation formulas suggest DAX measures → PBIT
        agg_formulas = sum(
            1 for f in r.fields
            if f.field_type == "formula" and
               re.search(r"\b(Sum|Count|Average|Maximum|Minimum)\s*\(", f.formula_text,
                         re.IGNORECASE)
        )
        if agg_formulas:
            scores["aggregation_formulas"] = -min(agg_formulas * 2, 8)

        # Single simple datasource
        sql_cmds = sum(1 for t in r.tables if t.sql_command.strip())
        if len(r.tables) == 1 and not sql_cmds:
            scores["single_simple_source"] = -5

        total = sum(scores.values())

        recommendation = "rdl" if total >= self.RDL_THRESHOLD else "pbit"

        reasons = []
        if recommendation == "rdl":
            if sr_count:
                reasons.append(f"{sr_count} subreport(s) → drill-through pages in RDL")
            if xtab:
                reasons.append(f"{xtab} cross-tab(s) → tablix in RDL")
            if pg_sections:
                reasons.append("Page headers/footers → RDL page bands")
            if g_count >= 3:
                reasons.append(f"{g_count} group levels → RDL grouping rows")
            if layout_objs:
                reasons.append(f"{layout_objs} line/box objects → pixel-perfect RDL")
        else:
            if has_detail_only:
                reasons.append("Simple detail-only tabular layout → Power BI table visual")
            if agg_formulas:
                reasons.append(f"{agg_formulas} aggregation formula(s) → DAX Measures in PBIT")
            if g_count in (1, 2):
                reasons.append(f"{g_count} group(s) → Matrix visual in PBIT")

        r.recommended_output = recommendation
        r.routing_reason     = "; ".join(reasons) or ("Score=%d" % total)
        r.routing_scores     = {**scores, "__total__": total}

        emit(f"  Routing: {recommendation.upper()}  (score={total})  {r.routing_reason}")
        return recommendation


# ═══════════════════════════════════════════════════════════════════════════════
# RDL BUILDER — generates Power BI Paginated Report (.rdl) XML
# Maps Crystal sections to SSRS/RDL report items
# ═══════════════════════════════════════════════════════════════════════════════
class RDLBuilder:
    """
    Crystal → RDL section mapping
    ─────────────────────────────────────────────────────────
    reportHeader  → ReportHeader band
    pageHeader    → PageHeader band
    groupHeader   → Tablix RowGroup header row
    detail        → Tablix detail row
    groupFooter   → Tablix RowGroup footer row (subtotals)
    pageFooter    → PageFooter band
    reportFooter  → ReportFooter band  (no direct RDL equiv → body footer)
    subreport     → SubReport item in body
    ─────────────────────────────────────────────────────────
    """

    # twips to cm conversion (1 inch = 1440 twips = 2.54 cm)
    TWIPS_TO_CM = 2.54 / 1440

    def build(self, report: CrystalReport, out_path: str):
        emit(f"  Building RDL  → {Path(out_path).name}")
        xml = self._render(report)
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(xml)
        emit(f"  ✓ RDL  saved  → {out_path}")

    def _cm(self, twips: int) -> str:
        """Convert twips to a CSS cm string."""
        return f"{max(twips * self.TWIPS_TO_CM, 0.1):.3f}cm"

    def _render(self, r: CrystalReport) -> str:
        tr    = FormulaTranslator()
        tname = r.tables[0].name if r.tables else "DataSet1"
        ds_name = "DataSet1"

        # ── Connection string ──────────────────────────────────────────────────
        if r.tables:
            t0 = r.tables[0]
            if t0.db_type in ("sql", "sqlserver", "native"):
                conn_type = "SQL"
                conn_str  = f"Data Source={t0.server};Initial Catalog={t0.database}"
            elif t0.db_type == "access":
                conn_type = "OLEDB"
                mdb = t0.database or t0.connection_string
                conn_str  = f"Provider=Microsoft.ACE.OLEDB.16.0;Data Source={mdb};"
            else:
                conn_type = "ODBC"
                conn_str  = t0.connection_string or f"DSN={t0.server}"
            sql = t0.sql_command.strip() or f"SELECT * FROM [{tname}]"
        else:
            conn_type = "ODBC"; conn_str = "DSN=MyDSN"; sql = "SELECT 1"

        db_fields = [f for t in r.tables for f in t.fields if f.field_type == "database"]
        col_width = "3cm"
        groups    = r.groups

        def x(s): return self._xml(str(s or ""))

        # ── Cell builders ──────────────────────────────────────────────────────
        def tablix_cols():
            return "\n".join(
                f"          <TablixColumn><Width>{col_width}</Width></TablixColumn>"
                for _ in db_fields)

        def header_cells():
            return "\n".join(
                f"""          <TablixCell><CellContents>
            <Textbox Name="hdr_{f.name}">
              <Paragraphs><Paragraph><TextRuns><TextRun>
                <Value>{x(f.name)}</Value>
                <Style><FontWeight>Bold</FontWeight></Style>
              </TextRun></TextRuns></Paragraph></Paragraphs>
              <Style><BackgroundColor>#E0E0E0</BackgroundColor></Style>
            </Textbox>
          </CellContents></TablixCell>""" for f in db_fields)

        def detail_cells():
            return "\n".join(
                f"""          <TablixCell><CellContents>
            <Textbox Name="det_{f.name}">
              <Paragraphs><Paragraph><TextRuns><TextRun>
                <Value>=Fields!{x(f.name)}.Value</Value>
              </TextRun></TextRuns></Paragraph></Paragraphs>
            </Textbox>
          </CellContents></TablixCell>""" for f in db_fields)

        def group_header_cells(g):
            gf = g["field"].split(".")[-1]
            cells = []
            for i, f in enumerate(db_fields):
                if i == 0:
                    cells.append(
                        f"""          <TablixCell><CellContents>
            <Textbox Name="gh_{gf}_{f.name}">
              <Paragraphs><Paragraph><TextRuns><TextRun>
                <Value>=Fields!{x(gf)}.Value</Value>
                <Style><FontWeight>Bold</FontWeight></Style>
              </TextRun></TextRuns></Paragraph></Paragraphs>
              <Style><BackgroundColor>#F0F0F0</BackgroundColor></Style>
            </Textbox>
          </CellContents></TablixCell>""")
                else:
                    cells.append("          <TablixCell />")
            return "\n".join(cells)

        def group_footer_cells(g):
            gf = g["field"].split(".")[-1]
            cells = []
            for f in db_fields:
                val = f"=Sum(Fields!{x(f.name)}.Value)" if f.data_type in ("number","currency") else "=&quot;Subtotal&quot;"
                cells.append(
                    f"""          <TablixCell><CellContents>
            <Textbox Name="gf_{gf}_{f.name}">
              <Paragraphs><Paragraph><TextRuns><TextRun>
                <Value>{val}</Value>
                <Style><FontWeight>Bold</FontWeight></Style>
              </TextRun></TextRuns></Paragraph></Paragraphs>
              <Style><BackgroundColor>#DDEEFF</BackgroundColor></Style>
            </Textbox>
          </CellContents></TablixCell>""")
            return "\n".join(cells)

        # ── TablixRows ─────────────────────────────────────────────────────────
        def tablix_rows():
            rows = [f"""        <TablixRow>
          <Height>0.6cm</Height>
          <TablixCells>
{header_cells()}
          </TablixCells>
        </TablixRow>"""]
            for g in groups:
                rows.append(f"""        <TablixRow>
          <Height>0.6cm</Height>
          <TablixCells>
{group_header_cells(g)}
          </TablixCells>
        </TablixRow>""")
            rows.append(f"""        <TablixRow>
          <Height>0.5cm</Height>
          <TablixCells>
{detail_cells()}
          </TablixCells>
        </TablixRow>""")
            for g in reversed(groups):
                rows.append(f"""        <TablixRow>
          <Height>0.6cm</Height>
          <TablixCells>
{group_footer_cells(g)}
          </TablixCells>
        </TablixRow>""")
            return "\n".join(rows)

        # ── TablixRowHierarchy ─────────────────────────────────────────────────
        # From working Invoice.rdl: groups use <Group> inside <TablixMember>,
        # NOT <TablixRowGroup>. TablixMembers only accepts TablixMember children.
        def row_hierarchy():
            parts = []
            # Header member
            parts.append("""        <TablixMember>
          <KeepWithGroup>After</KeepWithGroup>
          <RepeatOnNewPage>true</RepeatOnNewPage>
        </TablixMember>""")
            # Open one TablixMember per group (they nest)
            for g in groups:
                gf    = g["field"].split(".")[-1]
                gname = re.sub(r"[^A-Za-z0-9_]", "_", g["field"])
                sdir  = "Descending" if g.get("order") == "desc" else "Ascending"
                parts.append(f"""        <TablixMember>
          <Group Name="{gname}">
            <GroupExpressions>
              <GroupExpression>=Fields!{x(gf)}.Value</GroupExpression>
            </GroupExpressions>
          </Group>
          <SortExpressions>
            <SortExpression>
              <Value>=Fields!{x(gf)}.Value</Value>
              <Direction>{sdir}</Direction>
            </SortExpression>
          </SortExpressions>
          <TablixMembers>""")
            # Detail (innermost)
            parts.append("""            <TablixMember>
              <Group Name="Detail" />
              <TablixMembers><TablixMember /></TablixMembers>
            </TablixMember>""")
            # Close group TablixMembers in reverse
            for _ in groups:
                parts.append("""          </TablixMembers>
        </TablixMember>""")
            # Footer placeholder per group
            for _ in groups:
                parts.append("        <TablixMember />")
            return "\n".join(parts)

        # ── Parameters ─────────────────────────────────────────────────────────
        def params_xml():
            if not r.parameters: return ""
            _tm = {"string":"String","number":"Integer","currency":"Float",
                   "date":"DateTime","datetime":"DateTime","boolean":"Boolean"}
            pxml = []
            for p in r.parameters:
                ptype = _tm.get(p.get("type","string"), "String")
                dvals = ""
                if p.get("defaults"):
                    dvals = "<DefaultValues>" + "".join(
                        f"<Value>{x(str(d))}</Value>" for d in p["defaults"][:5]
                    ) + "</DefaultValues>"
                pxml.append(f"""  <ReportParameter Name="{x(p['name'])}">
    <DataType>{ptype}</DataType>
    <Prompt>{x(p.get('prompt', p['name']))}</Prompt>
    {dvals}
  </ReportParameter>""")
            return "<ReportParameters>\n" + "\n".join(pxml) + "\n</ReportParameters>"

        # ── Section helpers ────────────────────────────────────────────────────
        def section_items(stype):
            secs = [s for s in r.sections if s.section_type == stype]
            if not secs: return ""
            items = []; y = 0.2
            for sec in secs:
                for obj in sec.objects:
                    items.append(self._rdl_item(obj, y, tr, tname))
                    y += max(obj.get("height", 720) * self.TWIPS_TO_CM, 0.5)
            return "\n".join(items)

        def band_h(stype):
            secs = [s for s in r.sections if s.section_type == stype]
            return "1.5cm" if not secs else self._cm(sum(s.height or 1440 for s in secs))

        # ── Subreports ─────────────────────────────────────────────────────────
        def subreport_items():
            items = []; y = 0.5
            for sr in r.subreports:
                ptags = ""
                if sr.link_fields:
                    ptags = "<Parameters>" + "".join(
                        f"""<SubreportParameter Name="{x(lf['sub_field'])}">
              <Value>=Fields!{x(lf['main_field'])}.Value</Value>
            </SubreportParameter>"""
                        for lf in sr.link_fields if lf.get("main_field") and lf.get("sub_field")
                    ) + "</Parameters>"
                items.append(f"""        <Subreport Name="sr_{x(sr.name)}">
          <ReportName>{x(sr.name)}</ReportName>
          {ptags}
          <Top>{y:.2f}cm</Top><Left>0.5cm</Left>
          <Width>{max(sr.w*self.TWIPS_TO_CM,10):.2f}cm</Width>
          <Height>{max(sr.h*self.TWIPS_TO_CM,2):.2f}cm</Height>
        </Subreport>""")
                y += max(sr.h*self.TWIPS_TO_CM, 2) + 0.3
            return "\n".join(items)

        # ── Final structure from Invoice.rdl ───────────────────────────────────
        # Report > ReportSections > ReportSection > Body / Width / Page
        # Page > PageHeader / PageFooter / PageHeight / PageWidth / Margins
        has_pg_hdr  = bool(section_items("pageHeader").strip())
        has_pg_ftr  = bool(section_items("pageFooter").strip())
        has_rpt_hdr = bool(section_items("reportHeader").strip())
        has_rpt_ftr = bool(section_items("reportFooter").strip())

        pg_hdr_xml = f"""        <PageHeader>
          <Height>{band_h("pageHeader")}</Height>
          <PrintOnFirstPage>true</PrintOnFirstPage>
          <PrintOnLastPage>true</PrintOnLastPage>
          <ReportItems>{section_items("pageHeader")}</ReportItems>
        </PageHeader>""" if has_pg_hdr else ""

        pg_ftr_xml = f"""        <PageFooter>
          <Height>{band_h("pageFooter")}</Height>
          <PrintOnFirstPage>true</PrintOnFirstPage>
          <PrintOnLastPage>true</PrintOnLastPage>
          <ReportItems>{section_items("pageFooter")}</ReportItems>
        </PageFooter>""" if has_pg_ftr else ""

        rpt_hdr_xml = section_items("reportHeader") if has_rpt_hdr else (
            f"""        <Textbox Name="rpt_hdr">
          <Paragraphs><Paragraph><TextRuns><TextRun>
            <Value>{x(r.report_title)}</Value>
            <Style><FontSize>16pt</FontSize><FontWeight>Bold</FontWeight></Style>
          </TextRun></TextRuns></Paragraph></Paragraphs>
          <Top>0cm</Top><Left>0cm</Left><Width>18cm</Width><Height>1.2cm</Height>
        </Textbox>""")

        return f"""<?xml version="1.0" encoding="utf-8"?>
<Report MustUnderstand="df"
  xmlns="http://schemas.microsoft.com/sqlserver/reporting/2016/01/reportdefinition"
  xmlns:rd="http://schemas.microsoft.com/SQLServer/reporting/reportdesigner"
  xmlns:df="http://schemas.microsoft.com/sqlserver/reporting/2016/01/reportdefinition/defaultfontfamily">
  <rd:ReportUnitType>Cm</rd:ReportUnitType>
  <rd:ReportID>{str(uuid.uuid4())}</rd:ReportID>
  <df:DefaultFontFamily>Segoe UI</df:DefaultFontFamily>
  <AutoRefresh>0</AutoRefresh>
  <DataSources>
    <DataSource Name="{ds_name}">
      <ConnectionProperties>
        <DataProvider>{conn_type}</DataProvider>
        <ConnectString>{x(conn_str)}</ConnectString>
      </ConnectionProperties>
    </DataSource>
  </DataSources>
  <DataSets>
    <DataSet Name="{ds_name}">
      <Query>
        <DataSourceName>{ds_name}</DataSourceName>
        <CommandText>{x(sql)}</CommandText>
      </Query>
      <Fields>
{self._rdl_fields(db_fields)}
      </Fields>
    </DataSet>
  </DataSets>
  {params_xml()}
  <ReportSections>
    <ReportSection>
      <Body>
        <ReportItems>
{rpt_hdr_xml}
          <Tablix Name="MainTablix">
            <TablixBody>
              <TablixColumns>
{tablix_cols()}
              </TablixColumns>
              <TablixRows>
{tablix_rows()}
              </TablixRows>
            </TablixBody>
            <TablixColumnHierarchy>
              <TablixMembers>
{"".join("                <TablixMember />" for _ in db_fields)}
              </TablixMembers>
            </TablixColumnHierarchy>
            <TablixRowHierarchy>
              <TablixMembers>
{row_hierarchy()}
              </TablixMembers>
            </TablixRowHierarchy>
            <DataSetName>{ds_name}</DataSetName>
            <Top>1.5cm</Top><Left>0.5cm</Left><Width>18cm</Width>
          </Tablix>
{subreport_items()}
          {"" if not has_rpt_ftr else section_items("reportFooter")}
        </ReportItems>
        <Height>28cm</Height>
      </Body>
      <Width>21cm</Width>
      <Page>
{pg_hdr_xml}
{pg_ftr_xml}
        <PageHeight>29.7cm</PageHeight>
        <PageWidth>21cm</PageWidth>
        <LeftMargin>1.5cm</LeftMargin>
        <RightMargin>1.5cm</RightMargin>
        <TopMargin>1cm</TopMargin>
        <BottomMargin>1cm</BottomMargin>
      </Page>
    </ReportSection>
  </ReportSections>
  <ConsumeContainerWhitespace>true</ConsumeContainerWhitespace>
</Report>"""



    def _rdl_fields(self, fields: List[CrystalField]) -> str:
        lines = []
        for f in fields:
            _type_map = {"string":"String","number":"Integer","currency":"Float",
                         "date":"DateTime","datetime":"DateTime","boolean":"Boolean"}
            rtype = _type_map.get(f.data_type, "String")
            lines.append(
                f"""        <Field Name="{self._xml(f.name)}">
          <DataField>{self._xml(f.name)}</DataField>
          <rd:TypeName>{rtype}</rd:TypeName>
        </Field>"""
            )
        return "\n".join(lines)

    def _rdl_item(self, obj: dict, y_offset: float,
                  tr: "FormulaTranslator", tname: str) -> str:
        """Convert a Crystal report object dict to an RDL Textbox/Line/Rectangle."""
        left   = f"{obj.get('left', 0) * self.TWIPS_TO_CM:.3f}cm"
        top    = f"{y_offset:.3f}cm"
        width  = f"{max(obj.get('width', 1440) * self.TWIPS_TO_CM, 0.5):.3f}cm"
        height = f"{max(obj.get('height', 360) * self.TWIPS_TO_CM, 0.3):.3f}cm"
        nm     = re.sub(r"[^A-Za-z0-9_]", "_", obj.get("name", f"obj_{id(obj)}"))

        t = obj.get("type", "")
        if t == "text":
            text = self._xml(obj.get("text", ""))
            bold  = "<FontWeight>Bold</FontWeight>" if obj.get("bold") else ""
            italic= "<FontStyle>Italic</FontStyle>" if obj.get("italic") else ""
            fsize = f"<FontSize>{obj.get('font_size', 10)}pt</FontSize>"
            return f"""<Textbox Name="{nm}">
          <Paragraphs><Paragraph><TextRuns><TextRun>
            <Value>{text}</Value>
            <Style>{bold}{italic}{fsize}</Style>
          </TextRun></TextRuns></Paragraph></Paragraphs>
          <Top>{top}</Top><Left>{left}</Left>
          <Width>{width}</Width><Height>{height}</Height>
        </Textbox>"""
        elif t == "field":
            fname = re.sub(r"[^A-Za-z0-9_]", "_", obj.get("field_name", ""))
            return f"""<Textbox Name="{nm}">
          <Paragraphs><Paragraph><TextRuns><TextRun>
            <Value>=Fields!{fname}.Value</Value>
          </TextRun></TextRuns></Paragraph></Paragraphs>
          <Top>{top}</Top><Left>{left}</Left>
          <Width>{width}</Width><Height>{height}</Height>
        </Textbox>"""
        elif t == "line":
            return f"""<Line Name="{nm}">
          <Top>{top}</Top><Left>{left}</Left>
          <Width>{width}</Width><Height>0.05cm</Height>
        </Line>"""
        elif t == "box":
            fill = obj.get("fill_color", "")
            bg   = f"<BackgroundColor>{fill}</BackgroundColor>" if fill else ""
            return f"""<Rectangle Name="{nm}">
          <ReportItems />
          {bg}
          <Top>{top}</Top><Left>{left}</Left>
          <Width>{width}</Width><Height>{height}</Height>
        </Rectangle>"""
        else:
            return f"""<!-- {t} object '{nm}' skipped -->"""

    @staticmethod
    def _xml(s: str) -> str:
        """Escape a string for XML content."""
        return (str(s)
                .replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;").replace('"', "&quot;"))


# ── PBIT Builder ─────────────────────────────────────────────────────────────
# Fix summary (all 25 issues from reference diff):
# #1  Version='1.28' as UTF-16LE     #2  Table=DB table name (not report name)
# #3  Source columns populated        #4  cultures[] in model
# #5  Formula calc columns            #6  Real M query from connection info
# #7  layoutOptimization:0            #8  publicCustomVisuals:[]
# #9  Layout config v5.70+theme       #10 No 'filters' key on section
# #11 Section config={"objects":{}}   #12 textbox x=28.36 y=0 w=704.28 h=40.40
# #13 tableEx x=28.36 y=40.40 ...    #14 'query' SemanticQueryDataShapeCommand
# #15 'dataTransforms' full metadata  #16 projections Values=queryRefs
# #17 prototypeQuery Select=all cols  #18 RowColor hidden via SuppressedJoinPredicates
# #19 Settings Version:4              #20 Metadata Version:5
# #21 CreatedFromRelease              #22 ContentType="" for most entries
# #23 DiagramLayout with nodes        #24 SecurityBindings=empty bytes
# #25 Server-only (no CLI)
class PBITBuilder:
    TX=28.360413589364843; TY=0.0; TW=704.2836041358936; TH=40.399002493765586
    BX=28.360413589364843; BY=40.399002493765586; BW=704.2836041358936; BH=679.6009975062344
    PW=1280; PH=720
    # Full-name aliases used by the new multi-page / matrix methods
    TITLE_X = TX; TITLE_Y = TY; TITLE_W = TW; TITLE_H = TH
    TABLE_X = BX; TABLE_Y = BY; TABLE_W = BW; TABLE_H = BH
    def __init__(self): self.tr=FormulaTranslator()

    def build(self,r,out_path):
        emit(f"  Building PBIT \u2192 {Path(out_path).name}")
        pn=re.sub(r"[^A-Za-z0-9_]","_",r.report_name)
        mq={t.name:self._mq(t, r.record_selection_formula) for t in r.tables}

        # Diagnostic: log the M query source for each table so problems are visible
        for tname, expr in mq.items():
            src = "grafted" if any("BWPortfolio" in l or "dsn=" in l.lower() for l in expr) \
                  else "placeholder" if any("Placeholder" in l for l in expr) \
                  else "constructed"
            line2 = expr[1] if len(expr) > 1 else ""
            emit(f"  M query [{tname}] source={src}: {line2[:80]}")

        def _u16(obj, crlf=False, **kwargs):
            """Encode a JSON-serialisable object to UTF-16-LE bytes (no BOM).
            crlf=True: use \\r\\n line endings (DataModelSchema matches reference).
            crlf=False: use \\n line endings (all other JSON files)."""
            s = json.dumps(obj, ensure_ascii=False, **kwargs)
            if crlf:
                s = s.replace('\n', '\r\n')
            return s.encode("utf-16-le")

        with zipfile.ZipFile(out_path,"w",zipfile.ZIP_DEFLATED) as zf:
            # Version — UTF-16-LE, no BOM
            zf.writestr("Version", "1.28".encode("utf-16-le"))

            # Content_Types — UTF-8 WITH BOM (reference starts ef bb bf 3c)
            zf.writestr("[Content_Types].xml",
                        b"\xef\xbb\xbf" + self._ct().encode("utf-8"))

            # DataModelSchema — UTF-16-LE, CRLF line endings (matches reference)
            zf.writestr("DataModelSchema",
                        _u16(self._schema(r, mq), crlf=True, indent=2))

            # DiagramLayout — UTF-16-LE, LF line endings (matches reference)
            zf.writestr("DiagramLayout",
                        _u16(self._diag(r), indent=2))

            # Report/Layout — UTF-16-LE, no indent, LF
            zf.writestr("Report/Layout",
                        _u16(self._layout(r, pn), separators=(",", ":")))

            # Settings — UTF-16-LE, LF
            zf.writestr("Settings",
                        _u16(self._settings()))

            # Metadata — UTF-16-LE, LF
            zf.writestr("Metadata",
                        _u16(self._meta(r)))

            # SecurityBindings — DPAPI blob header (PBI regenerates on first save)
            zf.writestr("SecurityBindings", b"\x01\x00\x00\x00")

        emit(f"  \u2713 PBIT saved \u2192 {out_path}")

    @staticmethod
    def _extract_mdb_path(tbl) -> str:
        """
        Return the full .mdb/.accdb path from SDK connection info.
        Checks tbl.database, tbl.connection_string, tbl.server in order.
        Returns "" if not found.
        """
        import re as _re
        for raw in [getattr(tbl, "database", ""),
                    tbl.connection_string,
                    getattr(tbl, "server", "")]:
            raw = str(raw or "").strip()
            m = _re.search(r'(?:database=)(.+?)(?:;|$)', raw, _re.IGNORECASE)
            if m:
                raw = m.group(1).strip()
            if ".mdb" in raw.lower() or ".accdb" in raw.lower():
                return raw.replace("\\\\", "\\")
        return ""

    def _mq(self, tbl, record_selection: str = "") -> list:
        """
        Build the Power Query M expression for a table partition.
        Optionally appends a filter step translated from the Crystal record
        selection formula.
        """
        # Grafted _m_expr takes priority (set when seed_pbit explicitly supplied)
        if hasattr(tbl, "_m_expr") and tbl._m_expr:
            return tbl._m_expr

        db  = tbl.db_type.lower()
        c   = tbl.connection_string or ""

        def _add_filter(lines: list, filter_m: str) -> list:
            """Append a Table.SelectRows filter step if filter_m is non-empty."""
            if not filter_m or not filter_m.strip():
                return lines
            in_idx = next((i for i,l in enumerate(lines) if l.strip() == "in"), len(lines)-1)
            # Add trailing comma to the previous data step
            prev = lines[in_idx - 1]
            if not prev.rstrip().endswith(","):
                lines = lines[:in_idx-1] + [prev + ","] + lines[in_idx:]
                in_idx += 0  # in_idx stays same since we replaced not inserted
            # Derive last step variable name
            prev_step = lines[in_idx - 1].strip().split("=")[0].strip().rstrip(",")
            step_name = f"{tbl.name}_Filtered"
            filter_line = (f'    {step_name} = Table.SelectRows({prev_step}, '
                          f'each {filter_m})')
            return lines[:in_idx] + [filter_line, "in", f"    {step_name}"]

        # Translate Crystal record selection formula to M if present
        filter_m = ""
        if record_selection:
            try:
                filter_m = FormulaTranslator().to_m_filter(record_selection)
            except Exception:
                pass

        # ── Access / .mdb / .accdb ────────────────────────────────────────────────
        mdb_path = self._extract_mdb_path(tbl)
        if mdb_path or db == "access":
            if mdb_path:
                # Use OleDb.DataSource with Microsoft.ACE.OLEDB.16.0 (64-bit).
                # Reasons we cannot use alternatives:
                #   - Odbc.DataSource("dsn=...") fails: Crystal creates 32-bit DSNs,
                #     Power BI Desktop is 64-bit → architecture mismatch error
                #   - Access.Database(File.Contents("path")) → InvalidDataSourceReference
                #     in PBIT templates (File.Contents evaluated immediately at load)
                #   - Access.Database("path string") → "cannot convert to Binary"
                # OleDb.DataSource with ACE OLEDB 16.0:
                #   - 64-bit, installed with Office 2016/365
                #   - Takes file path directly in connection string
                #   - Treated as deferred data source by PBI (prompts user on first open)
                #   - No DSN required
                tbl_step = f'{tbl.name}_Table'
                conn = (f"Provider=Microsoft.ACE.OLEDB.16.0;"
                        f"Data Source={mdb_path};")
                lines = [
                    "let",
                    f'    Source = OleDb.DataSource("{conn}", []),',
                    f'    {tbl_step} = Source{{[Name="{tbl.name}"]}}[Data]',
                    "in",
                    f"    {tbl_step}",
                ]
                return _add_filter(lines, filter_m)
            emit(f"  WARN: db_type=access but no .mdb path found for '{tbl.name}'", "WARN")

        # ── SQL Server ────────────────────────────────────────────────────────────
        if db in ("sql", "sqlserver", "native"):
            srv = getattr(tbl, "server",   "") or ""
            db_ = getattr(tbl, "database", "") or ""
            for p in c.split(";"):
                pl = p.lower()
                if "server="   in pl: srv = srv or p.split("=", 1)[-1].strip()
                if "database=" in pl: db_ = db_ or p.split("=", 1)[-1].strip()
            if srv or db_:
                return [
                    "let",
                    f'    Source = Sql.Database("{srv}", "{db_}"),',
                    f'    {tbl.name}_Table = Source{{[Schema="dbo",Item="{tbl.name}"]}}[Data]',
                    "in",
                    f"    {tbl.name}_Table",
                ]

        # ── ODBC with explicit DSN in connection string or ServerName field ──────
        if db == "odbc" or "dsn=" in c.lower():
            # Try connection_string for explicit dsn=name
            dsn = ""
            for p in c.split(";"):
                if "dsn=" in p.lower():
                    dsn = p.split("=", 1)[-1].strip()
            # Fall back to ServerName — Crystal stores the DSN name there for ODBC
            if not dsn:
                dsn = getattr(tbl, "server", "") or ""
            if dsn:
                return [
                    "let",
                    f'    Source = Odbc.DataSource("dsn={dsn}", [HierarchicalNavigation=true]),',
                    f'    {tbl.name}_Table = Source{{[Name="{tbl.name}",Kind="Table"]}}[Data]',
                    "in",
                    f"    {tbl.name}_Table",
                ]

        # ── Unknown — placeholder ─────────────────────────────────────────────────
        emit(f"  WARN: cannot build M query for '{tbl.name}' "
             f"(db_type={tbl.db_type!r}, database={getattr(tbl,'database','')!r}, "
             f"conn={c!r})", "WARN")
        return [
            "let",
            f'    // TODO: configure data source for {tbl.name}',
            "    Source = #table(",
            "        type table [id = Int64.Type, name = text],",
            '        {{1, "Configure data source in Power BI Desktop"}}',
            "    )",
            "in",
            "    Source",
        ]

    def _schema(self,r,mq):                                             # Fix #4 cultures
        model = {
            "culture":"en-US",
            "dataAccessOptions":{"legacyRedirects":True,"returnErrorValuesAsNull":True},
            "defaultPowerBIDataSourceVersion":"powerBI_V3",
            "sourceQueryCulture":"en-US",
            "tables":[self._tbl(t,r,mq.get(t.name,[])) for t in r.tables],
            "cultures":[{"name":"en-US","linguisticMetadata":{"content":{"Version":"1.0.0","Language":"en-US"},"contentType":"json"}}],
            "annotations":[{"name":"PBI_QueryOrder","value":json.dumps([t.name for t in r.tables])},
                           {"name":"__PBI_TimeIntelligenceEnabled","value":"1"}],
        }
        # Add Crystal parameters as Power BI M parameters in the model expressions
        pbi_params = self._build_parameters(r)
        if pbi_params:
            model["expressions"] = pbi_params
        return {"name":str(uuid.uuid4()),"compatibilityLevel":1600,"model":model}

    def _build_parameters(self, r: "CrystalReport") -> list:
        """Convert Crystal Report parameters to PBI model expressions (M parameters)."""
        exprs = []
        _type_map = {
            "string":   "text",
            "number":   "number",
            "currency": "number",
            "date":     "date",
            "datetime": "datetime",
            "boolean":  "logical",
        }
        for p in r.parameters:
            pname = re.sub(r"[^A-Za-z0-9_ ]", "", p.get("name","")).strip()
            if not pname:
                continue
            ptype = _type_map.get(p.get("type","string"), "text")
            # Use first default value if available, otherwise a sensible placeholder
            defaults = p.get("defaults", [])
            if defaults:
                default_val = defaults[0]
                if ptype == "text":
                    expr_val = f'"{default_val}"'
                else:
                    expr_val = str(default_val)
            else:
                expr_val = '""' if ptype == "text" else "0"

            exprs.append({
                "name":        pname,
                "kind":        "m",
                "expression":  expr_val,
                "lineageTag":  str(uuid.uuid4()),
                "annotations": [
                    {"name": "PBI_ResultType",         "value": "Text"},
                    {"name": "PBI_NavigationStepName", "value": "Navigation"},
                ],
            })
            emit(f"  PBIT: parameter '{pname}' ({ptype}) → M expression {expr_val!r}")
        return exprs

    def _tbl(self,tbl,r,mq):
        """Build a table definition with columns, measures, and partitions."""
        cols    = []   # regular + calculated columns
        measures= []   # DAX measures (aggregating formulas)
        tr = self.tr

        # ── Regular database columns ──────────────────────────────────────────
        for f in tbl.fields:
            if f.field_type != "database": continue
            c = {
                "name":        f.name,
                "dataType":    self._dt(f.data_type),
                "sourceColumn":f.source_column,
                "lineageTag":  str(uuid.uuid4()),
                "summarizeBy": "sum" if f.data_type in("number","currency") else "none",
                "annotations": [{"name":"SummarizationSetBy","value":"Automatic"}],
            }
            if f.data_type in("number","currency"):
                c["formatString"] = "0"
            cols.append(c)

        # ── Formula fields → DAX measures or calculated columns ───────────────
        # Aggregating formulas (Sum/Count/Avg etc.) → measures
        # Row-level formulas → calculated columns
        agg_pattern = re.compile(
            r'\b(Sum|Count|Average|Maximum|Minimum|StdDev|Variance|DistinctCount)\s*\(',
            re.IGNORECASE)
        for f in r.fields:
            if f.field_type != "formula":
                continue
            fname = f.name.lstrip("{@").rstrip("}")
            dax   = tr.to_dax(f.formula_text, tbl.name)
            if not dax:
                dax = f'"{fname}"'

            if agg_pattern.search(f.formula_text or ""):
                # Aggregating → DAX Measure
                measures.append({
                    "name":        fname,
                    "expression":  dax,
                    "dataType":    self._dt(f.data_type),
                    "lineageTag":  str(uuid.uuid4()),
                    "formatString": "0" if f.data_type in("number","currency") else "",
                    "annotations": [{"name":"SummarizationSetBy","value":"Automatic"}],
                })
            else:
                # Row-level → Calculated Column
                cols.append({
                    "type":               "calculated",
                    "name":               fname,
                    "dataType":           self._dt(f.data_type),
                    "isDataTypeInferred": True,
                    "expression":         dax,
                    "lineageTag":         str(uuid.uuid4()),
                    "summarizeBy":        "none",
                    "annotations":        [{"name":"SummarizationSetBy","value":"Automatic"}],
                })

        # ── RowColor alternating row highlight column ─────────────────────────
        cols.append(self._rowcolor(tbl))

        # ── Assemble table definition ─────────────────────────────────────────
        tbl_def = {
            "name":       tbl.name,
            "lineageTag": str(uuid.uuid4()),
            "columns":    cols,
            "partitions": [{"name": tbl.name, "mode": "import",
                "source": {"type":"m","expression":
                    mq or ["let","    Source = #table({{}})","in","    Source"]}}],
            "annotations":[{"name":"PBI_NavigationStepName","value":"Navigation"},
                           {"name":"PBI_ResultType","value":"Table"}],
        }
        if measures:
            tbl_def["measures"] = measures
        return tbl_def

    def _rowcolor(self,tbl):
        id_col=next((f.name for f in tbl.fields if f.data_type in("number","currency")),"id")
        return {"type":"calculated","name":"RowColor","dataType":"string","isDataTypeInferred":True,
            "lineageTag":str(uuid.uuid4()),"summarizeBy":"none",
            "annotations":[{"name":"SummarizationSetBy","value":"Automatic"}],
            "expression":["","",f"VAR CurrentValue = '{tbl.name}'[{id_col}]","Var RecordNumber = ",
                "COUNTROWS(","    FILTER(",f"        ALL('{tbl.name}'),",
                f"        '{tbl.name}'[{id_col}] <= CurrentValue","    )",")",
                "","RETURN","    IF(MOD(RecordNumber, 2) = 0, \"#FFCCCC\", \"#FFFFFF\")"]}

    def _layout(self,r,pn):
        # Main page + one drill-through page per subreport
        sections = [self._section(r, pn, ordinal=0)]
        for idx, sr in enumerate(r.subreports, 1):
            sections.append(self._subreport_page(sr, r, ordinal=idx))

        return {"id":0,"reportId":str(uuid.uuid4()),
            "resourcePackages":[{"resourcePackage":{"name":"SharedResources","type":2,"disabled":False,
                "items":[{"type":202,"path":"BaseThemes/CY26SU02.json","name":"CY26SU02"}]}}],
            "sections": sections,
            "config":json.dumps({"version":"5.70",
                "themeCollection":{"baseTheme":{"name":"CY26SU02","type":2,
                    "version":{"visual":"2.6.0","report":"3.1.0","page":"2.3.0"}}},
                "activeSectionIndex":0,"defaultDrillFilterOtherVisuals":True,"linguisticSchemaSyncVersion":2,
                "settings":{"useNewFilterPaneExperience":True,"useStylableVisualContainerHeader":True,
                    "useEnhancedTooltips":True,"allowChangeFilterTypes":True,
                    "queryLimitOption":6,"exportDataMode":1,"useDefaultAggregateDisplayName":True},
                "objects":{"section":[{"properties":{"verticalAlignment":{"expr":{"Literal":{"Value":"'Top'"}}}}}],
                    "outspacePane":[{"properties":{"expanded":{"expr":{"Literal":{"Value":"false"}}}}}]}
            },separators=(",",":")),
            "layoutOptimization":0,
            "publicCustomVisuals":[]}

    def _section(self, r, pn, ordinal=0):
        title = r.report_title or r.report_name
        has_groups = len(r.groups) > 0

        # Page header objects → additional textboxes above the main table
        header_vcs = self._page_header_vcs(r)

        # Main visual: Matrix (grouped) or Table (flat)
        main_vc = self._matrix_vc(r) if has_groups else self._table_vc(r)

        # Page footer objects
        footer_vcs = self._page_footer_vcs(r)

        all_vcs = [self._title_vc(title)] + header_vcs + [main_vc] + footer_vcs

        return {"name":pn,"displayName":pn,"ordinal":ordinal,
            "visualContainers": all_vcs,
            "width":self.PW,"height":self.PH,
            "config":json.dumps({"objects":{}},separators=(",",":"))}

    def _subreport_page(self, sr: "CrystalSubreport", r: CrystalReport, ordinal: int) -> dict:
        """Drill-through page for a subreport, with its own table visual."""
        page_name = re.sub(r"[^A-Za-z0-9_]", "_", sr.name) or f"Subreport_{ordinal}"
        display   = sr.name or f"Subreport {ordinal}"

        # Fake a mini-report with just the link fields as columns
        if sr.link_fields:
            db_flds = [
                CrystalField(lf["sub_field"], "database", "string",
                             table_name=r.tables[0].name if r.tables else "Table")
                for lf in sr.link_fields
                if lf.get("sub_field")
            ]
        else:
            db_flds = [f for t in r.tables for f in t.fields
                       if f.field_type == "database"][:6]

        tbl_vc = self._build_table_vc_from_fields(
            db_flds,
            r.tables[0].name if r.tables else "Table",
            y_offset=self.TABLE_Y,
        )

        # Info textbox
        nm   = uuid.uuid4().hex[:20]
        info_cfg = {"name":nm,"layouts":[{"id":0,"position":{
                "x":self.TITLE_X,"y":self.TITLE_Y,"z":0,
                "width":self.TITLE_W,"height":self.TITLE_H,"tabOrder":1000}}],
            "singleVisual":{"visualType":"textbox","objects":{"general":[{"properties":{"paragraphs":[{
                "textRuns":[{"value":f"Drill-through: {display}",
                             "textStyle":{"fontWeight":"bold","fontSize":"14pt"}}],
                "horizontalTextAlignment":"center"}]}}]}}}
        info_vc = {"x":self.TITLE_X,"y":self.TITLE_Y,"z":0,
                   "width":self.TITLE_W,"height":self.TITLE_H,
                   "config":json.dumps(info_cfg,separators=(",",":")),
                   "filters":"[]","tabOrder":1000}

        return {
            "name":        page_name,
            "displayName": display,
            "ordinal":     ordinal,
            "visualContainers": [info_vc, tbl_vc],
            "width":  self.PW, "height": self.PH,
            "config": json.dumps({"objects":{},"drillthrough":{"target":True}},
                                 separators=(",",":")),
        }

    def _page_header_vcs(self, r: CrystalReport) -> list:
        """Turn Crystal pageHeader text objects into PBI textbox visuals."""
        vcs = []
        y = self.TITLE_H + 2  # just below title
        for sec in r.sections:
            if sec.section_type != "pageHeader": continue
            for obj in sec.objects:
                if obj.get("type") not in ("text", "field_heading"): continue
                text = obj.get("text", "")
                if not text.strip(): continue
                nm = uuid.uuid4().hex[:20]
                w  = min(self.TABLE_W, max(obj.get("width",0) * 0.0176, 100))
                h  = max(obj.get("height",0) * 0.0176, 20)
                cfg = {"name":nm,"layouts":[{"id":0,"position":{
                        "x":self.TITLE_X,"y":y,"z":1,"width":w,"height":h,"tabOrder":500}}],
                    "singleVisual":{"visualType":"textbox","objects":{"general":[{"properties":{"paragraphs":[{
                        "textRuns":[{"value":text,"textStyle":{"fontWeight":"bold","fontSize":"10pt"}}]}]}}]}}}
                vcs.append({"x":self.TITLE_X,"y":y,"z":1,"width":w,"height":h,
                            "config":json.dumps(cfg,separators=(",",":")),"filters":"[]","tabOrder":500})
                y += h + 2
        return vcs

    def _page_footer_vcs(self, r: CrystalReport) -> list:
        """Turn Crystal pageFooter text objects into PBI textbox visuals near bottom."""
        vcs = []
        y = self.PH - 60
        for sec in r.sections:
            if sec.section_type != "pageFooter": continue
            for obj in sec.objects:
                if obj.get("type") not in ("text", "field"): continue
                text = obj.get("text", obj.get("field_name", ""))
                if not text.strip(): continue
                nm = uuid.uuid4().hex[:20]
                w  = min(self.TABLE_W, max(obj.get("width",0) * 0.0176, 100))
                h  = max(obj.get("height",0) * 0.0176, 18)
                cfg = {"name":nm,"layouts":[{"id":0,"position":{
                        "x":self.TITLE_X,"y":y,"z":1,"width":w,"height":h,"tabOrder":200}}],
                    "singleVisual":{"visualType":"textbox","objects":{"general":[{"properties":{"paragraphs":[{
                        "textRuns":[{"value":text,"textStyle":{"fontSize":"9pt"}}]}]}}]}}}
                vcs.append({"x":self.TITLE_X,"y":y,"z":1,"width":w,"height":h,
                            "config":json.dumps(cfg,separators=(",",":")),"filters":"[]","tabOrder":200})
                y += h + 2
        return vcs

    def _matrix_vc(self, r: CrystalReport) -> dict:
        """
        Build a Matrix visual for grouped reports.
        Maps Crystal Group Header/Footer → Matrix row groups with subtotals.
        Crystal Detail → Matrix detail rows.
        """
        if not r.tables or not r.tables[0].fields: return self._pvc()
        tbl      = r.tables[0]
        entity   = tbl.name
        alias    = entity[0].lower()
        db_flds  = [f for f in tbl.fields if f.field_type == "database"]
        if not db_flds: return self._pvc()

        group_fields = [g["field"].split(".")[-1] for g in r.groups]
        # Value fields = db fields that are NOT group fields
        val_flds = [f for f in db_flds if f.name not in group_fields]
        if not val_flds:
            val_flds = db_flds  # fallback

        def qn(f):
            return f"Sum({entity}.{f.name})" if f.data_type in ("number","currency") \
                   else f"{entity}.{f.name}"

        # Rows = group fields; Columns = value fields
        row_qnames = [f"{entity}.{gf}" for gf in group_fields if gf]
        val_qnames = [qn(f) for f in val_flds]

        projections = {
            "Rows":   [{"queryRef": q} for q in row_qnames],
            "Values": [{"queryRef": q} for q in val_qnames],
        }

        proto_sel = []
        for gf in group_fields:
            proto_sel.append({
                "Column": {"Expression": {"SourceRef": {"Source": alias}},
                           "Property": gf},
                "Name": f"{entity}.{gf}",
                "NativeReferenceName": gf,
            })
        for f in val_flds:
            if f.data_type in ("number","currency"):
                proto_sel.append({
                    "Aggregation": {
                        "Expression": {"Column": {
                            "Expression": {"SourceRef": {"Source": alias}},
                            "Property": f.name}},
                        "Function": 0},  # Sum
                    "Name": qn(f),
                    "NativeReferenceName": f.name,
                })
            else:
                proto_sel.append({
                    "Column": {"Expression": {"SourceRef": {"Source": alias}},
                               "Property": f.name},
                    "Name": qn(f),
                    "NativeReferenceName": f.name,
                })

        # Sort by first group field
        sort_prop  = group_fields[0] if group_fields else (db_flds[0].name)
        proto_q = {
            "Version": 2,
            "From":    [{"Name": alias, "Entity": entity, "Type": 0}],
            "Select":  proto_sel,
            "OrderBy": [{"Direction": 1, "Expression": {"Column": {
                "Expression": {"SourceRef": {"Source": alias}},
                "Property":   sort_prop,
            }}}],
        }

        nm = uuid.uuid4().hex[:20]
        cfg = {
            "name": nm,
            "layouts": [{"id":0,"position":{
                "x":self.TABLE_X,"y":self.TABLE_Y,"z":11001,
                "width":self.TABLE_W,"height":self.TABLE_H,"tabOrder":12001}}],
            "singleVisual": {
                "visualType":             "matrix",
                "projections":            projections,
                "prototypeQuery":          proto_q,
                "drillFilterOtherVisuals": True,
                "objects": {
                    "subTotals": [{"properties": {
                        "rowSubtotalsPosition": {"expr": {"Literal": {"Value": "'Bottom'"}}}
                    }}],
                },
            },
        }
        return {
            "x":self.TABLE_X,"y":self.TABLE_Y,"z":11001,
            "width":self.TABLE_W,"height":self.TABLE_H,
            "config":json.dumps(cfg,separators=(",",":")),
            "filters":"[]","tabOrder":12001,
        }

    def _build_table_vc_from_fields(self, db_flds, entity, y_offset=None):
        """Build a tableEx VC directly from a list of CrystalField objects."""
        if y_offset is None:
            y_offset = self.TABLE_Y
        if not db_flds:
            return self._pvc()
        alias = entity[0].lower()

        def qn(f):
            return f"Sum({entity}.{f.name})" if f.data_type in ("number","currency") \
                   else f"{entity}.{f.name}"

        col_qnames  = [qn(f) for f in db_flds]
        n_vis       = len(db_flds)
        sort_col    = next((f.name for f in db_flds if f.data_type == "string"),
                           db_flds[0].name)
        projections = [{"queryRef": q} for q in col_qnames]

        def ce(f):
            return {"Column": {"Expression": {"SourceRef": {"Source": alias}},
                               "Property": f.name}}

        proto_sel = [{**ce(f), "Name": qn(f), "NativeReferenceName": f.name}
                     for f in db_flds]
        proto_q   = {
            "Version": 2,
            "From":    [{"Name": alias, "Entity": entity, "Type": 0}],
            "Select":  proto_sel,
            "OrderBy": [{"Direction": 1, "Expression": {"Column": {
                "Expression": {"SourceRef": {"Source": alias}},
                "Property": sort_col,
            }}}],
        }
        rc_agg = {"Aggregation": {
            "Expression": {"Column": {
                "Expression": {"SourceRef": {"Source": alias}},
                "Property": "RowColor"}},
            "Function": 3}}
        sem_sel = list(proto_sel) + [{**rc_agg, "Name": f"Min({entity}.RowColor)"}]
        sem_q   = {"Commands": [{"SemanticQueryDataShapeCommand": {
            "Query": {"Version": 2,
                      "From":   [{"Name": alias, "Entity": entity, "Type": 0}],
                      "Select": sem_sel,
                      "OrderBy": proto_q["OrderBy"]},
            "Binding": {
                "Primary": {"Groupings": [{"Projections": list(range(n_vis+1)), "Subtotal": 1}]},
                "DataReduction": {"DataVolume": 3, "Primary": {"Window": {"Count": 500}}},
                "SuppressedJoinPredicates": [n_vis], "Version": 1},
            "ExecutionMetricsKind": 1,
        }}]}

        rc_cf = {"Aggregation": {
            "Expression": {"Column": {
                "Expression": {"SourceRef": {"Entity": entity}},
                "Property": "RowColor"}},
            "Function": 3}}
        cf_objects = [{"properties": {"backColor": {"solid": {"color": {"expr": rc_cf}}}},
                       "selector": {"data": [{"dataViewWildcard": {"matchingOption": 1}}],
                                    "metadata": q}}
                      for q in col_qnames]

        dt = {
            "objects":            {"values": cf_objects},
            "projectionOrdering": {"Values": list(range(n_vis))},
            "queryMetadata": {
                "Select": [{"Restatement": f.name, "Name": qn(f),
                            "Type": 3 if f.data_type in ("number","currency") else 2048,
                            **({"Format":"0"} if f.data_type in ("number","currency") else {})}
                           for f, q in zip(db_flds, col_qnames)],
                "Filters": [{"type": 2 if f.data_type in ("number","currency") else 0,
                             "expression": {"Column": {"Expression": {
                                 "SourceRef": {"Entity": entity}},
                                 "Property": f.name}}}
                            for f in db_flds],
            },
            "visualElements": [{"DataRoles": [
                {"Name":"Values","Projection":i,"isActive":False} for i in range(n_vis)
            ]}],
            "selects": self._build_selects(db_flds, entity, col_qnames, sort_col),
        }

        nm  = uuid.uuid4().hex[:20]
        cfg = {"name": nm,
               "layouts": [{"id":0,"position":{
                   "x":self.TABLE_X,"y":y_offset,"z":11001,
                   "width":self.TABLE_W,"height":self.TABLE_H,"tabOrder":12001}}],
               "singleVisual": {
                   "visualType":             "tableEx",
                   "projections":            {"Values": projections},
                   "prototypeQuery":          proto_q,
                   "drillFilterOtherVisuals": True,
                   "objects":                {"values": cf_objects},
               }}
        return {
            "x": self.TABLE_X,"y": y_offset,"z": 11001,
            "width": self.TABLE_W,"height": self.TABLE_H,
            "config":         json.dumps(cfg, separators=(",",":")),
            "filters":        "[]","tabOrder": 12001,
            "query":          json.dumps(sem_q,  separators=(",",":")),
            "dataTransforms": json.dumps(dt,     separators=(",",":")),
        }

    @staticmethod
    def _build_selects(db_flds, entity, col_qnames, sort_col):
        """Build the selects[] array for dataTransforms — one entry per visible col + hidden RowColor."""
        selects = []
        for f, qn in zip(db_flds, col_qnames):
            e: dict = {
                "displayName": f.name,
                "queryName":   qn,
                "roles":       {"Values": True},
                "type": {
                    "category":      None,
                    "underlyingType": 260 if f.data_type in ("number","currency") else 1,
                },
                "expr": {"Column": {
                    "Expression": {"SourceRef": {"Entity": entity}},
                    "Property":   f.name,
                }},
            }
            if f.data_type in ("number","currency"):
                e["format"] = "0"
            if f.name == sort_col:
                e["sort"]      = 1
                e["sortOrder"] = 0
            selects.append(e)
        # Hidden RowColor select with relatedObjects back-references
        rc_agg = {"Aggregation": {
            "Expression": {"Column": {
                "Expression": {"SourceRef": {"Entity": entity}},
                "Property": "RowColor",
            }},
            "Function": 3,
        }}
        selects.append({
            "displayName": "First RowColor",
            "queryName":   f"Min({entity}.RowColor)",
            "roles":       {},
            "type":        {"category": None, "underlyingType": 1},
            "expr":        rc_agg,
            "relatedObjects": {
                "values": {
                    "backColor": [
                        {"data": [{"dataViewWildcard": {"matchingOption": 1}}],
                         "metadata": qn}
                        for qn in col_qnames
                    ]
                }
            },
        })
        return selects

    def _title_vc(self,title):
        nm=uuid.uuid4().hex[:20]
        cfg={"name":nm,"layouts":[{"id":0,"position":{"x":self.TX,"y":self.TY,"z":0,"width":self.TW,"height":self.TH,"tabOrder":1000}}],
            "singleVisual":{"visualType":"textbox","objects":{"general":[{"properties":{"paragraphs":[{
                "textRuns":[{"value":title,"textStyle":{"fontWeight":"bold","fontSize":"16pt"}}],
                "horizontalTextAlignment":"center"}]}}]}}}
        return {"x":self.TX,"y":self.TY,"z":0,"width":self.TW,"height":self.TH,  # Fix #12
            "config":json.dumps(cfg,separators=(",",":")),"filters":"[]","tabOrder":1000}

    def _table_vc(self, r):
        """Flat tableEx visual — used when the report has no groups."""
        if not r.tables or not r.tables[0].fields: return self._pvc()
        tbl     = r.tables[0]
        db_flds = [f for f in tbl.fields if f.field_type == "database"]
        if not db_flds: return self._pvc()
        return self._build_table_vc_from_fields(db_flds, tbl.name)

    def _pvc(self):
        nm=uuid.uuid4().hex[:20]
        cfg={"name":nm,"layouts":[{"id":0,"position":{"x":self.BX,"y":self.BY,"z":11001,"width":self.BW,"height":self.BH,"tabOrder":12001}}],
            "singleVisual":{"visualType":"textbox","objects":{"general":[{"properties":{"paragraphs":[{
                "textRuns":[{"value":"\u26a0 No columns extracted. Run on Windows with SAP Crystal Runtime.","textStyle":{"fontSize":"12pt"}}]}]}}]}}}
        return {"x":self.BX,"y":self.BY,"z":11001,"width":self.BW,"height":self.BH,
            "config":json.dumps(cfg,separators=(",",":")),"filters":"[]","tabOrder":12001}

    @staticmethod
    def _ct():                                                         # Fix #22
        return ('<?xml version="1.0" encoding="utf-8"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="json" ContentType="" />'
                '<Override PartName="/Version" ContentType="" />'
                '<Override PartName="/DataModelSchema" ContentType="" />'
                '<Override PartName="/DiagramLayout" ContentType="" />'
                '<Override PartName="/Report/Layout" ContentType="" />'
                '<Override PartName="/Settings" ContentType="application/json" />'
                '<Override PartName="/Metadata" ContentType="application/json" />'
                '<Override PartName="/SecurityBindings" ContentType="" />'
                '</Types>')

    @staticmethod
    def _diag(r):                                                      # Fix #23
        return {"version":"1.1.0","diagrams":[{"ordinal":0,"scrollPosition":{"x":0,"y":0},
            "nodes":[{"location":{"x":80+i*320,"y":90},"nodeIndex":t.name,"size":{"height":300,"width":284},"zIndex":i+1}
                     for i,t in enumerate(r.tables)],
            "name":f"{r.report_name} Data Model","zoomValue":100,"pinKeyFieldsTo":"TopEdge"}]}

    @staticmethod
    def _settings():                                                   # Fix #19
        return {"Version":4,"ReportSettings":{},"QueriesSettings":{"TypeDetectionEnabled":True,"RelationshipImportEnabled":True,"Version":"2.152.856.0"}}

    @staticmethod
    def _meta(r):                                                      # Fix #20,21
        return {"Version":5,"AutoCreatedRelationships":[],"CreatedFrom":"Crystal2PBI",
                "CreatedFromRelease":datetime.now().strftime("%Y.%m"),"ReportSource":Path(r.file_path).name}

    @staticmethod
    def _dt(t):
        return {"string":"string","text":"string","number":"int64","numeric":"int64",
                "currency":"double","decimal":"double","date":"dateTime","datetime":"dateTime",
                "time":"dateTime","boolean":"boolean"}.get(t.lower(),"string")

# ── Excel Documenter ─────────────────────────────────────────────────────────
class ExcelDocumenter:
    H  = PatternFill("solid",fgColor="1F4E79")
    HF = Font(name="Calibri",bold=True,color="FFFFFF",size=11)
    TF = Font(name="Calibri",bold=True,size=14,color="1F4E79")
    AF = PatternFill("solid",fgColor="D9E1F2")
    BR = Border(bottom=Side(style="thin",color="B8CCE4"),
                right =Side(style="thin",color="B8CCE4"))

    def document(self, r, out):
        emit(f"  Documenting → {Path(out).name}")
        wb = openpyxl.Workbook()
        self._summary(wb, r)
        self._datasources(wb, r)
        self._sql_commands(wb, r)
        self._fields(wb, r)
        self._formulas(wb, r)
        self._params(wb, r)
        self._groups_sorts(wb, r)
        self._sections(wb, r)
        self._subreports(wb, r)
        self._routing(wb, r)
        self._complexity(wb, r)
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        wb.save(out)
        emit(f"  ✓ Excel saved → {out}")

    def _summary(self, wb, r):
        ws = wb.create_sheet("Summary")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        ws["A1"] = "Crystal Reports Documentation"
        ws["A1"].font = self.TF
        rows = [
            ("Report Name",          r.report_name),
            ("Report Title",         r.report_title),
            ("Subject",              r.subject),
            ("Author",               r.author),
            ("Source File",          r.file_path),
            ("Parse Method",         r.parse_method),
            ("Recommended Output",   r.recommended_output.upper() or "—"),
            ("Routing Reason",       r.routing_reason or "—"),
            ("Tables",               len(r.tables)),
            ("DB Fields",            sum(1 for f in r.fields if f.field_type=="database")),
            ("Formula Fields",       sum(1 for f in r.fields if f.field_type=="formula")),
            ("Parameters",           len(r.parameters)),
            ("Groups",               len(r.groups)),
            ("Sort Fields",          len(r.sort_fields)),
            ("Subreports",           len(r.subreports)),
            ("Has Record Selection", bool(r.record_selection_formula)),
            ("Documentation Date",   datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for i, (lbl, val) in enumerate(rows, 3):
            ws.cell(i, 1, lbl).font = Font(bold=True)
            ws.cell(i, 2, str(val))
            if i % 2 == 0:
                ws.cell(i, 1).fill = self.AF
                ws.cell(i, 2).fill = self.AF
        if r.warnings:
            r0 = len(rows) + 4
            ws.cell(r0, 1, "⚠ Warnings").font = Font(bold=True, color="FF0000")
            for j, w in enumerate(r.warnings):
                ws.cell(r0 + 1 + j, 1, w)

    def _datasources(self, wb, r):
        ws = wb.create_sheet("DataSources")
        self._hdr(ws, ["Table","Alias","DB Type","Connection Type",
                        "Server","Database","Connection String","Fields"])
        for i, t in enumerate(r.tables, 2):
            self._row(ws, i, [t.name, t.alias, t.db_type, t.connection_type,
                               t.server, t.database, t.connection_string, len(t.fields)])
        self._af(ws)

    def _sql_commands(self, wb, r):
        ws = wb.create_sheet("SQL Commands")
        self._hdr(ws, ["Table","Has SQL Command","SQL / Command Text"])
        ws.column_dimensions["C"].width = 80
        for i, t in enumerate(r.tables, 2):
            has_sql = "Yes" if t.sql_command.strip() else "No"
            self._row(ws, i, [t.name, has_sql, t.sql_command[:500]])
        # Record selection formula
        if r.record_selection_formula:
            row = len(r.tables) + 3
            ws.cell(row, 1, "Record Selection Formula").font = Font(bold=True)
            ws.cell(row, 3, r.record_selection_formula[:500])
        if r.group_selection_formula:
            row = len(r.tables) + 4
            ws.cell(row, 1, "Group Selection Formula").font = Font(bold=True)
            ws.cell(row, 3, r.group_selection_formula[:500])
        self._af(ws)

    def _fields(self, wb, r):
        ws = wb.create_sheet("Fields")
        self._hdr(ws, ["Name","Type","DataType","Table","Source Column","Formula (first 200)"])
        for i, f in enumerate(r.fields, 2):
            self._row(ws, i, [f.name, f.field_type, f.data_type,
                               f.table_name, f.source_column, f.formula_text[:200]])
        self._af(ws)

    def _formulas(self, wb, r):
        ws = wb.create_sheet("Formulas")
        self._hdr(ws, ["Name","DataType","Crystal Formula","DAX Translation"])
        tr = FormulaTranslator()
        tn = r.tables[0].name if r.tables else "Table"
        for i, f in enumerate([x for x in r.fields if x.field_type == "formula"], 2):
            self._row(ws, i, [f.name, f.data_type, f.formula_text,
                               tr.to_dax(f.formula_text, tn)])
        ws.column_dimensions["C"].width = 70
        ws.column_dimensions["D"].width = 70

    def _params(self, wb, r):
        ws = wb.create_sheet("Parameters")
        self._hdr(ws, ["Name","Type","Prompt","Multi-value","Range","Defaults","PBI Equivalent"])
        for i, p in enumerate(r.parameters, 2):
            self._row(ws, i, [
                p["name"], p.get("type",""), p.get("prompt",""),
                "Yes" if p.get("allow_multiple") else "No",
                "Yes" if p.get("allow_range") else "No",
                ", ".join(p.get("defaults",[])),
                f"Power BI Parameter: {p['name']} ({p.get('type','')})",
            ])
        self._af(ws)

    def _groups_sorts(self, wb, r):
        ws = wb.create_sheet("Groups & Sorting")
        self._hdr(ws, ["Level","Type","Field","Direction","PBI Mapping"])
        row = 2
        for g in r.groups:
            self._row(ws, row, [
                g.get("group_level", ""), "Group",
                g["field"], g.get("order","asc"),
                "Matrix RowGroup / DAX GROUPBY",
            ])
            row += 1
        for sf in r.sort_fields:
            self._row(ws, row, [
                "", "Sort",
                sf.field_name, sf.direction,
                "Power BI Sort on visual / ORDERBY in DAX",
            ])
            row += 1
        self._af(ws)

    def _sections(self, wb, r):
        ws = wb.create_sheet("Sections")
        self._hdr(ws, ["Name","Type","Suppress","Height(twips)","BG Color",
                        "Objects","PBI Mapping"])
        MAP = {
            "reportHeader": "Top of report page / title textbox",
            "pageHeader":   "Visual column headers / page header textbox",
            "groupHeader":  "Matrix RowGroup header row",
            "detail":       "Table/Matrix detail rows",
            "groupFooter":  "Matrix subtotal row",
            "pageFooter":   "Page footer textbox (limited in PBIT)",
            "reportFooter": "Grand totals / summary visuals",
        }
        for i, s in enumerate(r.sections, 2):
            self._row(ws, i, [
                s.name, s.section_type,
                "Yes" if s.suppress else "No",
                s.height, s.background_color,
                len(s.objects),
                MAP.get(s.section_type, "Custom"),
            ])
        self._af(ws)
        # Objects detail sub-table
        ws2 = wb.create_sheet("Section Objects")
        self._hdr(ws2, ["Section","Object Name","Type","Left","Top",
                         "Width","Height","Text/Field","Font","Bold"])
        row = 2
        for sec in r.sections:
            for obj in sec.objects:
                self._row(ws2, row, [
                    sec.name, obj.get("name",""), obj.get("type",""),
                    obj.get("left",0), obj.get("top",0),
                    obj.get("width",0), obj.get("height",0),
                    obj.get("text", obj.get("field_name", obj.get("chart_title",""))),
                    obj.get("font_name",""),
                    "Yes" if obj.get("bold") else "",
                ])
                row += 1
        self._af(ws2)

    def _subreports(self, wb, r):
        ws = wb.create_sheet("Subreports")
        self._hdr(ws, ["Subreport Name","Parent Section","Link Fields",
                        "PBI Equivalent","Notes"])
        for i, sr in enumerate(r.subreports, 2):
            links = "; ".join(
                f"{lf.get('main_field','')}→{lf.get('sub_field','')}"
                for lf in sr.link_fields
            )
            self._row(ws, i, [
                sr.name, sr.section,
                links or "(no links)",
                "Drill-through page in Power BI",
                "Create a separate PBIT page named '" + sr.name + "' with drill-through filters set",
            ])
        self._af(ws)

    def _routing(self, wb, r):
        ws = wb.create_sheet("Output Routing")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 55
        ws["A1"] = "Output Format Routing Decision"
        ws["A1"].font = self.TF

        rec = (r.recommended_output or "pbit").upper()
        color = "0070C0" if rec == "PBIT" else "7030A0"
        ws.cell(3, 1, "Recommended Output").font = Font(bold=True)
        ws.cell(3, 2, rec).font = Font(bold=True, color=color)

        ws.cell(4, 1, "Routing Reason").font = Font(bold=True)
        ws.cell(4, 2, r.routing_reason or "—")

        ws.cell(6, 1, "Score Breakdown").font = Font(bold=True)
        self._hdr_at(ws, 7, ["Factor","Points","Direction"])
        row = 8
        for k, v in r.routing_scores.items():
            if k == "__total__": continue
            direction = "→ RDL" if v > 0 else "→ PBIT"
            self._row(ws, row, [k.replace("_"," ").title(), v, direction])
            row += 1
        total = r.routing_scores.get("__total__", 0)
        ws.cell(row, 1, "TOTAL SCORE").font = Font(bold=True)
        ws.cell(row, 2, total).font = Font(bold=True)
        ws.cell(row, 3, f"Threshold = {ReportAnalyzer.RDL_THRESHOLD}  (≥{ReportAnalyzer.RDL_THRESHOLD} → RDL)").font = Font(italic=True)

        row += 2
        ws.cell(row, 1, "Crystal → PBI Section Mapping").font = Font(bold=True, size=12)
        row += 1
        self._hdr_at(ws, row, ["Crystal Section","Power BI Equivalent"])
        mapping = [
            ("Report Header",  "Top of report page / title textbox"),
            ("Page Header",    "Visual column headers / header textbox"),
            ("Group Header",   "Matrix RowGroup header / group label"),
            ("Details",        "Table or Matrix detail rows"),
            ("Group Footer",   "Matrix subtotal row / DAX CALCULATE totals"),
            ("Page Footer",    "Page footer textbox (limited in PBIT)"),
            ("Subreport",      "Drill-through page or Tooltip page"),
            ("Report Footer",  "Grand total row / summary card visuals"),
        ]
        for r2, (c, p) in enumerate(mapping, row+1):
            self._row(ws, r2, [c, p])

    def _complexity(self, wb, r):
        ws = wb.create_sheet("Complexity")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 55
        ws["A1"] = "Migration Complexity Analysis"
        ws["A1"].font = self.TF
        ff = sum(1 for f in r.fields if f.field_type == "formula")
        sc = (min(len(r.fields)*0.5,20) + min(ff*2,30) + len(r.parameters)*3
              + len(r.groups)*2 + len(r.subreports)*10
              + (5 if r.record_selection_formula else 0)
              + max(len(r.tables)-1,0)*3)
        cx = "Low" if sc<20 else "Medium" if sc<50 else "High" if sc<80 else "Very High"
        items = [
            ("Tables",             len(r.tables),       ""),
            ("Total Fields",       len(r.fields),       ""),
            ("Formula Fields",     ff,                  "Each needs DAX translation"),
            ("Parameters",         len(r.parameters),   "Requires PBI parameter setup"),
            ("Group Levels",       len(r.groups),       "Maps to Matrix RowGroup"),
            ("Sort Fields",        len(r.sort_fields),  "Maps to visual Sort"),
            ("Subreports",         len(r.subreports),   "Each = drill-through page in PBI"),
            ("Record Selection",   int(bool(r.record_selection_formula)), "Maps to M filter"),
            ("Complexity Score",   f"{sc:.0f}/100",     ""),
            ("Overall Complexity", cx,                   "Low<20|Med<50|High<80|VHigh"),
        ]
        for i, (lbl, val, note) in enumerate(items, 3):
            ws.cell(i, 1, lbl).font = Font(bold=True)
            ws.cell(i, 2, str(val))
            ws.cell(i, 3, note)
            if i % 2 == 0:
                for c in range(1, 4):
                    ws.cell(i, c).fill = self.AF
        cm = {"Low":"70AD47","Medium":"FFC000","High":"FF7043","Very High":"C00000"}
        ws.cell(len(items)+3, 2).fill = PatternFill("solid", fgColor=cm.get(cx,"FFFFFF"))

    # ── helpers ──────────────────────────────────────────────────────────────
    def _hdr(self, ws, headers, start_row=1):
        for j, h in enumerate(headers, 1):
            c = ws.cell(start_row, j, h)
            c.fill = self.H; c.font = self.HF
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = f"A{start_row+1}"

    def _hdr_at(self, ws, row, headers):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row, j, h)
            c.fill = self.H; c.font = self.HF

    def _row(self, ws, row, vals):
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, str(v) if v is not None else "")
            c.border = self.BR
            if row % 2 == 0:
                c.fill = self.AF

    @staticmethod
    def _af(ws):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 60)

# ── Migration Engine ─────────────────────────────────────────────────────────
class MigrationEngine:
    def __init__(self):
        self.parser   = RPTParser()
        self.analyzer = ReportAnalyzer()
        self.builder  = PBITBuilder()
        self.rdl      = RDLBuilder()
        self.doc      = ExcelDocumenter()

    def run(self, rpt_paths, output_dir, mode="both", seed_pbit=None,
            force_output: str = ""):
        """
        force_output: "" (auto-route) | "pbit" | "rdl"
        """
        results = {"processed": [], "errors": [], "output_dir": output_dir}
        os.makedirs(output_dir, exist_ok=True)

        for rpt in rpt_paths:
            try:
                emit(f"\n{'='*52}")
                emit(f"Processing: {Path(rpt).name}")
                r    = self.parser.parse(rpt, seed_pbit=seed_pbit)
                base = Path(rpt).stem
                res  = {"file": rpt, "warnings": r.warnings,
                        "parse_method": r.parse_method}

                # Always analyze to set recommended_output and scores
                output_fmt = force_output or self.analyzer.analyze(r)

                if mode in ("document", "both"):
                    dp = os.path.join(output_dir, f"{base}_documentation.xlsx")
                    self.doc.document(r, dp)
                    res["excel_doc"] = dp

                if mode in ("convert", "both"):
                    if output_fmt == "rdl":
                        op = os.path.join(output_dir, f"{base}.rdl")
                        self.rdl.build(r, op)
                        res["rdl"]    = op
                        res["output"] = "rdl"
                    else:
                        op = os.path.join(output_dir, f"{base}.pbit")
                        self.builder.build(r, op)
                        res["pbit"]   = op
                        res["output"] = "pbit"

                    res["routing_score"]  = r.routing_scores.get("__total__", 0)
                    res["routing_reason"] = r.routing_reason

                results["processed"].append(res)
                emit(f"✓ Done: {base}  [{r.parse_method}]  → {output_fmt.upper()}")

            except Exception as e:
                emit(f"{Path(rpt).name}: {e}\n{traceback.format_exc()}", "ERROR")
                results["errors"].append({"file": rpt, "error": str(e)})

        return results

# ── Flask Server (server-only, Fix #25) ──────────────────────────────────────
app    = Flask(__name__, static_folder=".", static_url_path="")
engine = MigrationEngine()   # includes parser, analyzer, builder, rdl, doc

# In-memory registry: report_stem → seed schema dict
# Populated by /seed upload or auto-detected from co-located .pbit files
_seed_registry: Dict[str, dict] = {}
_seed_lock = threading.Lock()


@app.route("/")
def index(): return send_from_directory(".","index.html")


@app.route("/status")
def status():
    with _seed_lock:
        seeded = list(_seed_registry.keys())
    sdk_detail = {
        None:  "Not available",
        "clr": "SAP Crystal SDK for .NET via pythonnet (clr)",
        "com": "SAP Crystal Runtime via win32com (COM)",
    }.get(CRYSTAL_SDK_MODE, CRYSTAL_SDK_MODE)
    return jsonify({
        "version":        SERVER_VERSION,
        "platform":       platform.system(),
        "crystal_sdk":    CRYSTAL_SDK_AVAILABLE,
        "sdk_mode":       CRYSTAL_SDK_MODE,
        "sdk_detail":     sdk_detail,
        "server":         "running",
        "seeded_reports": seeded,
    })


@app.route("/seed", methods=["POST"])
def seed_endpoint():
    """
    Upload a reference .pbit file to seed the column schema for a report.

    Accepts multipart/form-data with field "file" (.pbit).
    Optionally also accepts field "report_stem" to associate the schema
    with a specific .rpt base name (defaults to the PBIT filename stem).

    The seeded schema is stored in _seed_registry and automatically used
    when converting the matching .rpt file.

    Returns: {"status":"ok","stem":str,"tables":[str...],"columns":int}
    """
    if "file" not in request.files:
        return jsonify({"status":"error","message":"No file field in request"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".pbit"):
        return jsonify({"status":"error","message":"File must be a .pbit"}), 400

    stem = request.form.get("report_stem", "") or Path(f.filename).stem
    # Remove common suffixes so "AlphaISOsByCountry-Reference" → "AlphaISOsByCountry"
    stem = re.sub(r"[-_]?[Rr]eference$","", stem)

    pbit_bytes = f.read()
    schema = PBITSeeder().extract_schema(pbit_bytes)
    if not schema:
        return jsonify({"status":"error","message":"Could not extract schema from PBIT"}), 422

    with _seed_lock:
        _seed_registry[stem] = schema

    table_names = [t["name"] for t in schema["tables"]]
    total_cols  = sum(len(t["columns"]) for t in schema["tables"])
    emit(f"Seed registered: stem={stem!r}  tables={table_names}  columns={total_cols}")
    return jsonify({
        "status":  "ok",
        "stem":    stem,
        "tables":  table_names,
        "columns": total_cols,
    })


@app.route("/seed_status")
def seed_status():
    """List all currently registered seed schemas."""
    with _seed_lock:
        info = {}
        for stem, schema in _seed_registry.items():
            info[stem] = {
                "tables":  [t["name"] for t in schema["tables"]],
                "columns": sum(len(t["columns"]) for t in schema["tables"]),
            }
    return jsonify(info)


@app.route("/diagnose")
def diagnose():
    """
    Scans the machine for SAP Crystal Reports .NET assemblies and registry
    entries.  Use this to find the correct CRYSTAL_ASSEMBLY_PATH value when
    the server starts in the wrong mode or the CLR load fails.

    Returns JSON with:
      sdk_mode        : "clr" | "com" | null
      env_override    : value of CRYSTAL_ASSEMBLY_PATH env var (if set)
      dll_found       : list of dirs that actually contain the Engine DLL
      all_searched    : every dir that was scanned (whether DLL found or not)
      registry_keys   : relevant SAP registry keys and their values
      gac_entries     : CrystalDecisions* entries in the GAC
      recommendation  : what to do next
    """
    result: dict = {
        "sdk_mode":     CRYSTAL_SDK_MODE,
        "env_override": os.environ.get("CRYSTAL_ASSEMBLY_PATH", ""),
        "dll_found":    [],
        "all_searched": [],
        "registry_keys": {},
        "gac_entries":  [],
        "recommendation": "",
    }

    if platform.system() != "Windows":
        result["recommendation"] = "Not running on Windows — SDK unavailable on this platform."
        return jsonify(result)

    # Search all candidate dirs
    bridge = CrystalSDKBridge()
    all_dirs = bridge._find_all_asm_dirs()
    result["all_searched"] = all_dirs
    for d in all_dirs:
        dll = os.path.join(d, "CrystalDecisions.CrystalReports.Engine.dll")
        if os.path.isfile(dll):
            result["dll_found"].append({
                "dir": d,
                "dll": dll,
                "also_has_shared": os.path.isfile(
                    os.path.join(d, "CrystalDecisions.Shared.dll")),
            })

    # Registry scan
    try:
        import winreg
        reg_paths = [
            (winreg.HKEY_LOCAL_MACHINE,
             r"SOFTWARE\SAP BusinessObjects\Crystal Reports for Visual Studio\Default"),
            (winreg.HKEY_LOCAL_MACHINE,
             r"SOFTWARE\SAP BusinessObjects\Crystal Reports for Visual Studio"),
            (winreg.HKEY_LOCAL_MACHINE,
             r"SOFTWARE\WOW6432Node\SAP BusinessObjects\Crystal Reports for Visual Studio"),
            (winreg.HKEY_LOCAL_MACHINE,
             r"SOFTWARE\SAP BusinessObjects\Suite XI 4.0\Crystal Reports"),
            (winreg.HKEY_LOCAL_MACHINE,
             r"SOFTWARE\Business Objects\Crystal Reports for Visual Studio\Default"),
        ]
        for hive, rp in reg_paths:
            try:
                with winreg.OpenKey(hive, rp) as k:
                    vals = {}
                    i = 0
                    while True:
                        try:
                            name, val, _ = winreg.EnumValue(k, i)
                            vals[name] = str(val)
                            i += 1
                        except OSError:
                            break
                    if vals:
                        result["registry_keys"][rp] = vals
            except (FileNotFoundError, OSError):
                pass
    except ImportError:
        result["registry_keys"]["error"] = "winreg not available"

    # GAC scan
    import glob
    gac = r"C:\Windows\Microsoft.NET\assembly\GAC_MSIL"
    if os.path.isdir(gac):
        result["gac_entries"] = [
            e for e in glob.glob(os.path.join(gac, "CrystalDecisions*"))
            if os.path.isdir(e)
        ]

    # Recommendation
    if result["dll_found"]:
        best = result["dll_found"][0]["dir"]
        result["recommendation"] = (
            f"DLL found. If the server still fails, set:\n"
            f'  set CRYSTAL_ASSEMBLY_PATH={best}\n'
            f"then restart the server."
        )
    elif result["gac_entries"]:
        result["recommendation"] = (
            "DLL not found on filesystem but GAC entries exist. "
            "The assembly may be GAC-only. Try: pip install pythonnet --upgrade "
            "and restart — pythonnet should resolve GAC assemblies automatically."
        )
    elif result["registry_keys"]:
        result["recommendation"] = (
            "Registry keys found but DLL location not identified. "
            "Check the registry values above and set CRYSTAL_ASSEMBLY_PATH "
            "to the directory containing CrystalDecisions.CrystalReports.Engine.dll."
        )
    else:
        result["recommendation"] = (
            "No Crystal Reports installation detected. "
            "Install SAP Crystal Reports, Developer Version for Visual Studio from: "
            "https://www.sap.com/products/technology-platform/crystal-reports.html  "
            "Then restart the server."
        )

    return jsonify(result)


@app.route("/document", methods=["POST"])
def route_doc():  return _job("document")

@app.route("/convert",  methods=["POST"])
def route_conv(): return _job("convert")

@app.route("/both",     methods=["POST"])
def route_both(): return _job("both")


def _job(mode: str) -> Response:
    global _sse_messages
    with _sse_lock: _sse_messages.clear()

    d   = request.json or {}
    inp = d.get("input_folder",  "./input")
    out = d.get("output_folder", "./output")
    explicit_seed = d.get("seed_pbit", "")
    # Caller can force "pbit" or "rdl"; default "" = auto-route
    force_output  = d.get("output_format", "").lower().strip()
    if force_output not in ("pbit", "rdl", ""):
        force_output = ""

    files_raw = sorted(Path(inp).glob("*.rpt")) + sorted(Path(inp).glob("*.RPT"))
    # Deduplicate by resolved absolute path (Windows filesystem is case-insensitive,
    # so *.rpt and *.RPT both match the same file)
    seen_paths: set = set()
    files = []
    for p in files_raw:
        key = str(p.resolve()).lower()
        if key not in seen_paths:
            seen_paths.add(key)
            files.append(str(p))
    if not files:
        return jsonify({"status":"error",
                        "message":f"No .rpt files in {inp}"}), 400

    def _worker():
        import time as _time
        for idx, rpt in enumerate(files):
            # OPTIMIZATION: Throttle CPU usage by adding small sleep between files
            if idx > 0:
                _time.sleep(0.5)  # Brief pause to let system cool down
            
            stem = Path(rpt).stem
            seed = explicit_seed or None
            try:
                emit(f"\n{'='*52}")
                emit(f"Processing: {Path(rpt).name} ({idx+1}/{len(files)})")
                r = engine.parser.parse(rpt, seed_pbit=seed)

                output_fmt = force_output or engine.analyzer.analyze(r)
                base = Path(rpt).stem

                if mode in ("document", "both"):
                    dp = os.path.join(out, f"{base}_documentation.xlsx")
                    engine.doc.document(r, dp)

                if mode in ("convert", "both"):
                    os.makedirs(out, exist_ok=True)
                    if output_fmt == "rdl":
                        engine.rdl.build(r, os.path.join(out, f"{base}.rdl"))
                    else:
                        engine.builder.build(r, os.path.join(out, f"{base}.pbit"))

                emit(f"✓ Done: {base} [{r.parse_method}] → {output_fmt.upper()}")
            except Exception as e:
                emit(f"{Path(rpt).name}: {e}\n{traceback.format_exc()}", "ERROR")

    os.makedirs(out, exist_ok=True)
    t = threading.Thread(target=_worker, daemon=True)
    t.start(); t.join(timeout=600)  # Increased timeout from 300 to 600s

    with _sse_lock: msgs = list(_sse_messages)
    return jsonify({"status":"done","mode":mode,"messages":msgs,
                    "files_processed":len(files),"output_folder":out,
                    "output_format": force_output or "auto"})


@app.route("/logs")
def logs_sse():
    def gen():
        seen = 0
        import time
        while True:
            with _sse_lock: new=_sse_messages[seen:]; seen=len(_sse_messages)
            for m in new: yield f"data: {m}\n\n"
            if not new: time.sleep(0.3)
    return Response(gen(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

if __name__=="__main__":
    ap=argparse.ArgumentParser(description="Crystal2PBI Web Server")
    ap.add_argument("--port",type=int,default=5000); ap.add_argument("--host",default="0.0.0.0"); ap.add_argument("--debug",action="store_true")
    args=ap.parse_args()
    os.makedirs("input",exist_ok=True); os.makedirs("output",exist_ok=True)
    _mode_label = {
        "clr": "SAP Crystal SDK for .NET  (pythonnet/clr)",
        "com": "SAP Crystal Runtime        (win32com/COM)",
        None:  "Heuristic OLE              (cross-platform, no SDK)",
    }[CRYSTAL_SDK_MODE]
    print(f"""
╔══════════════════════════════════════════════════════════╗
║         Crystal Reports → Power BI  Server               ║
╠══════════════════════════════════════════════════════════╣
║  Version : {SERVER_VERSION:<46s}║
║  URL     : http://localhost:{args.port:<29}║
║  Mode    : {_mode_label:<46s}║
║  Input   : ./input/      Output: ./output/               ║
╚══════════════════════════════════════════════════════════╝
Press Ctrl+C to stop.
""")
    app.run(host=args.host,port=args.port,debug=args.debug,use_reloader=False)
